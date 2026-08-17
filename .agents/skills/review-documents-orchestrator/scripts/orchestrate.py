#!/usr/bin/env python3
"""Deterministically prepare and finalize a design-document review run."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Callable


ROLE_DIRECTORIES = {
    "checklists": "input/checklists",
    "references": "input/references",
    "targets": "input/targets",
}
FORMAT_SKILLS = {
    ".xlsx": ("xlsx-document", "xlsx_document.py"),
    ".docx": ("docx-document", "docx_document.py"),
    ".pptx": ("pptx-document", "pptx_document.py"),
    ".pdf": ("pdf-document", "pdf_document.py"),
    ".png": ("image-document", "image_document.py"),
    ".jpg": ("image-document", "image_document.py"),
    ".jpeg": ("image-document", "image_document.py"),
    ".tif": ("image-document", "image_document.py"),
    ".tiff": ("image-document", "image_document.py"),
    ".bmp": ("image-document", "image_document.py"),
    ".webp": ("image-document", "image_document.py"),
}
LEGACY_FORMATS = {".xls": ".xlsx", ".doc": ".docx", ".ppt": ".pptx"}
SUPPORTED_EXTENSIONS = set(FORMAT_SKILLS) | set(LEGACY_FORMATS)
CHECKLIST_EXTENSIONS = {".xlsx", ".xls"}
RUN_ID_PATTERN = re.compile(r"^\d{12}$")
OWNER_MARKER = ".review-orchestrator-owner"


class OrchestrationError(RuntimeError):
    """Raised when a run cannot proceed without risking incomplete output."""


Runner = Callable[[list[str], Path], None]


def _skills_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _is_link_like(path: Path) -> bool:
    return path.is_symlink() or path.is_junction()


def _relative(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError as exc:
        raise OrchestrationError(f"リポジトリルート外のパスは扱えません: {path}") from exc


def _validated_repo_path(root: Path, path: Path, label: str) -> Path:
    """Return a lexical absolute path after boundary and link-like checks."""
    root = root.resolve()
    supplied = Path(path)
    if ".." in supplied.parts:
        raise OrchestrationError(f"{label}に`..`は使用できません: {path}")
    lexical = supplied if supplied.is_absolute() else root / supplied
    lexical = Path(os.path.abspath(lexical))
    try:
        relative = lexical.relative_to(root)
    except ValueError as exc:
        raise OrchestrationError(f"{label}はリポジトリルート配下にしてください: {path}") from exc
    current = root
    for part in relative.parts:
        current = current / part
        if _is_link_like(current):
            raise OrchestrationError(
                f"{label}の祖先または対象にsymlink/junctionは使用できません: "
                f"{current.relative_to(root).as_posix()}"
            )
    resolved = lexical.resolve(strict=False)
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise OrchestrationError(f"{label}がリポジトリルート外へ解決されます: {path}") from exc
    return lexical


def _manifest_relative_path(
    root: Path, value: Any, expected_prefix: str, label: str
) -> str:
    if not isinstance(value, str) or not value.strip():
        raise OrchestrationError(f"manifestの{label}がありません")
    supplied = Path(value)
    if supplied.is_absolute() or ".." in supplied.parts:
        raise OrchestrationError(f"manifestの{label}はroot相対パスにしてください: {value}")
    candidate = _validated_repo_path(root, root / supplied, f"manifestの{label}")
    relative = candidate.relative_to(root.resolve()).as_posix()
    prefix = expected_prefix.rstrip("/") + "/"
    if relative != expected_prefix.rstrip("/") and not relative.startswith(prefix):
        raise OrchestrationError(
            f"manifestの{label}は`{expected_prefix}/`配下にしてください: {value}"
        )
    return relative


def _read_json(path: Path) -> dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except FileNotFoundError as exc:
        raise OrchestrationError(f"ファイルが見つかりません: {path}") from exc
    except json.JSONDecodeError as exc:
        raise OrchestrationError(f"JSONを解析できません: {path} ({exc})") from exc
    if not isinstance(data, dict):
        raise OrchestrationError(f"JSONのルートはオブジェクトである必要があります: {path}")
    return data


def _write_json(root: Path, path: Path, data: dict[str, Any]) -> None:
    path = _validated_repo_path(root, path, "JSON生成先")
    if path.exists() or _is_link_like(path):
        raise OrchestrationError(f"既存JSONは上書きしません: {_relative(path, root)}")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("x", encoding="utf-8") as stream:
        stream.write(json.dumps(data, ensure_ascii=False, indent=2) + "\n")


def _run_command(command: list[str], cwd: Path) -> None:
    try:
        subprocess.run(command, cwd=cwd, check=True)
    except FileNotFoundError as exc:
        raise OrchestrationError(f"実行ファイルが見つかりません: {command[0]}") from exc
    except subprocess.CalledProcessError as exc:
        raise OrchestrationError(
            f"処理に失敗しました（終了コード {exc.returncode}）: {' '.join(command)}"
        ) from exc


def inventory_inputs(root: Path) -> dict[str, list[Path]]:
    root = root.resolve()
    inventory: dict[str, list[Path]] = {}
    for role, relative_directory in ROLE_DIRECTORIES.items():
        directory = root / relative_directory
        _validated_repo_path(root, directory, f"{relative_directory}入力フォルダ")
        if not directory.is_dir():
            raise OrchestrationError(f"入力フォルダがありません: {relative_directory}")
        files: list[Path] = []
        unsupported: list[str] = []
        for path in sorted(directory.rglob("*")):
            _validated_repo_path(root, path, "入力ファイルまたはフォルダ")
            if not path.is_file() or path.name.lower() == "readme.md" or path.name.startswith("."):
                continue
            if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append(path)
            else:
                unsupported.append(_relative(path, root))
        if unsupported:
            raise OrchestrationError(
                "未対応形式の入力があります:\n- " + "\n- ".join(unsupported)
            )
        inventory[role] = files
    if not inventory["checklists"]:
        raise OrchestrationError(
            "input/checklists/にチェックリストがありません。.xlsxファイルを配置してください。"
        )
    if not inventory["targets"]:
        raise OrchestrationError("input/targets/にレビュー対象ファイルがありません。")
    invalid_checklists = [
        _relative(path, root)
        for path in inventory["checklists"]
        if path.suffix.lower() not in CHECKLIST_EXTENSIONS
    ]
    if invalid_checklists:
        raise OrchestrationError(
            "結果列を追加するため、チェックリストは.xlsx（旧形式は.xls）にしてください:\n- "
            + "\n- ".join(invalid_checklists)
        )
    return inventory


def _validate_run_id(run_id: str) -> None:
    if not RUN_ID_PATTERN.fullmatch(run_id):
        raise OrchestrationError("run-idはローカル時刻のyyyyMMddhhmm（12桁）で指定してください")
    try:
        datetime.strptime(run_id, "%Y%m%d%H%M")
    except ValueError as exc:
        raise OrchestrationError(
            "run-idは実在するローカル日時のyyyyMMddhhmmで指定してください"
        ) from exc


def _run_artifact_paths(root: Path, run_id: str) -> list[Path]:
    return [
        root / "work" / "review-runs" / run_id,
        root / "work" / "markdown" / run_id,
        root / "work" / "images" / run_id,
        root / "work" / "converted-office" / run_id,
        root / "output" / "reviews" / run_id,
    ]


def _reject_existing_run_artifacts(root: Path, run_id: str) -> None:
    paths = _run_artifact_paths(root, run_id)
    for path in paths:
        _validated_repo_path(root, path, "run-id別生成先")
    conflicts = [path for path in paths if path.exists() or _is_link_like(path)]
    if conflicts:
        raise OrchestrationError(
            "同じrun-idのパスが既にあります。空でも再利用しません:\n- "
            + "\n- ".join(_relative(path, root) for path in conflicts)
        )


def _cleanup_owned_directory(path: Path, ownership_token: str) -> None:
    if not path.exists() and not _is_link_like(path):
        return
    if _is_link_like(path) or not path.is_dir():
        return
    marker = path / OWNER_MARKER
    try:
        if marker.read_text(encoding="utf-8") != ownership_token:
            return
    except OSError:
        return
    shutil.rmtree(path)


def _create_owned_work_directories(root: Path, run_id: str) -> tuple[list[Path], str]:
    paths = _run_artifact_paths(root, run_id)[:4]
    ownership_token = uuid.uuid4().hex
    created: list[Path] = []
    try:
        for path in paths:
            _validated_repo_path(root, path, "run-id別work生成先")
            path.mkdir(parents=True)
            try:
                (path / OWNER_MARKER).write_text(ownership_token, encoding="utf-8")
            except Exception:
                path.rmdir()
                raise
            created.append(path)
    except Exception:
        for path in reversed(created):
            _cleanup_owned_directory(path, ownership_token)
        raise
    return paths, ownership_token


def _legacy_destination(root: Path, run_id: str, role: str, source: Path) -> Path:
    role_root = root / ROLE_DIRECTORIES[role]
    relative = source.relative_to(role_root)
    destination = root / "work" / "converted-office" / run_id / role / relative
    return destination.with_suffix(LEGACY_FORMATS[source.suffix.lower()])


def _markdown_destination(root: Path, run_id: str, role: str, original: Path) -> Path:
    role_root = root / ROLE_DIRECTORIES[role]
    relative = original.relative_to(role_root)
    return (
        root / "work" / "markdown" / run_id / role / relative.parent / f"{relative.name}.md"
    )


def _images_destination(root: Path, run_id: str, role: str, original: Path) -> Path:
    role_root = root / ROLE_DIRECTORIES[role]
    relative = original.relative_to(role_root)
    source_directory = f"{relative.stem}_{relative.suffix.lstrip('.').lower()}"
    return root / "work" / "images" / run_id / role / relative.parent / source_directory


def _format_command(
    root: Path,
    run_id: str,
    role: str,
    source: Path,
    original: Path,
    markdown: Path,
) -> list[str]:
    suffix = source.suffix.lower()
    try:
        skill_name, script_name = FORMAT_SKILLS[suffix]
    except KeyError as exc:
        raise OrchestrationError(f"Markdown変換に未対応の形式です: {source}") from exc
    script = _skills_root() / skill_name / "scripts" / script_name
    images = _images_destination(root, run_id, role, original)
    images.mkdir(parents=True, exist_ok=True)
    command = [
        sys.executable,
        str(script),
        "to-markdown",
        str(source),
        str(markdown),
        "--role",
        role[:-1] if role.endswith("s") else role,
        "--repo-root",
        str(root),
        "--images-dir",
        str(images),
    ]
    if suffix in {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}:
        command.extend(["--ocr", "--lang", "jpn+eng+jpn_vert"])
    return command


def _rewrite_legacy_frontmatter(
    markdown: Path,
    original_path: str,
    intermediate_path: str,
    original_format: str,
    intermediate_format: str,
) -> None:
    try:
        lines = markdown.read_text(encoding="utf-8-sig").splitlines()
    except FileNotFoundError as exc:
        raise OrchestrationError(f"変換後Markdownがありません: {markdown}") from exc
    if not lines or lines[0].strip() != "---":
        raise OrchestrationError(f"MarkdownにYAML frontmatterがありません: {markdown}")
    try:
        end = lines.index("---", 1)
    except ValueError as exc:
        raise OrchestrationError(f"MarkdownのYAML frontmatterが閉じていません: {markdown}") from exc
    replacements = {
        "source_path": json.dumps(original_path, ensure_ascii=False),
        "source_name": json.dumps(Path(original_path).name, ensure_ascii=False),
        "source_format": json.dumps(original_format.lstrip("."), ensure_ascii=False),
    }
    frontmatter: list[str] = []
    seen: set[str] = set()
    for line in lines[1:end]:
        key = line.split(":", 1)[0].strip() if ":" in line else ""
        if key in {"intermediate_path", "intermediate_format"}:
            continue
        if key in replacements:
            frontmatter.append(f"{key}: {replacements[key]}")
            seen.add(key)
        else:
            frontmatter.append(line)
    for key in ("source_path", "source_name", "source_format"):
        if key not in seen:
            frontmatter.append(f"{key}: {replacements[key]}")
    frontmatter.extend(
        [
            f"intermediate_path: {json.dumps(intermediate_path, ensure_ascii=False)}",
            f"intermediate_format: {json.dumps(intermediate_format.lstrip('.'), ensure_ascii=False)}",
        ]
    )
    markdown.write_text(
        "\n".join(["---", *frontmatter, "---", *lines[end + 1 :]]).rstrip() + "\n",
        encoding="utf-8",
    )


def _ocr_embedded_images(
    root: Path,
    run_id: str,
    role: str,
    original: Path,
    parent_markdown: Path,
    images_directory: Path,
    runner: Runner,
) -> None:
    _validated_repo_path(root, parent_markdown, "親Markdown")
    _validated_repo_path(root, images_directory, "抽出画像フォルダ")
    supported_images = {
            ".png",
            ".jpg",
            ".jpeg",
            ".tif",
            ".tiff",
            ".bmp",
            ".webp",
    }
    all_files = sorted(path for path in images_directory.rglob("*") if path.is_file())
    if not all_files:
        return
    role_root = root / ROLE_DIRECTORIES[role]
    relative_source = original.relative_to(role_root)
    ocr_root = (
        root
        / "work"
        / "markdown"
        / run_id
        / "embedded-image-ocr"
        / role
        / relative_source.parent
        / f"{relative_source.stem}_{relative_source.suffix.lstrip('.').lower()}"
    )
    image_script = _skills_root() / "image-document" / "scripts" / "image_document.py"
    sections: list[str] = ["", "## 抽出画像OCR", ""]
    for image_path in all_files:
        _validated_repo_path(root, image_path, "抽出画像")
        relative_image = image_path.relative_to(images_directory)
        if image_path.suffix.lower() not in supported_images:
            sections.extend(
                [
                    f"### `{_relative(image_path, root)}`",
                    "",
                    f"- 元文書: `{_relative(original, root)}`",
                    f"- 抽出画像: `{_relative(image_path, root)}`",
                    f"- 拡張子: `{image_path.suffix or '(なし)'}`",
                    "- 判定上の扱い: OCR不可のため要確認。視覚情報を推測しない。",
                    "",
                ]
            )
            continue
        ocr_markdown = _validated_repo_path(
            root,
            ocr_root / relative_image.parent / f"{relative_image.name}.md",
            "抽出画像OCR Markdown生成先",
        )
        ocr_markdown.parent.mkdir(parents=True, exist_ok=True)
        runner(
            [
                sys.executable,
                str(image_script),
                "to-markdown",
                str(image_path),
                str(ocr_markdown),
                "--role",
                role[:-1] if role.endswith("s") else role,
                "--repo-root",
                str(root),
                "--ocr",
                "--lang",
                "jpn+eng+jpn_vert",
            ],
            root,
        )
        if not ocr_markdown.is_file():
            raise OrchestrationError(f"抽出画像OCRのMarkdownが作られませんでした: {ocr_markdown}")
        sections.extend(
            [
                f"### `{_relative(image_path, root)}`",
                "",
                f"- 元文書: `{_relative(original, root)}`",
                f"- 抽出画像: `{_relative(image_path, root)}`",
                f"- OCR Markdown: `{_relative(ocr_markdown, root)}`",
                "",
                ocr_markdown.read_text(encoding="utf-8-sig").rstrip(),
                "",
            ]
        )
    try:
        original_markdown = parent_markdown.read_text(encoding="utf-8-sig").rstrip()
    except FileNotFoundError as exc:
        raise OrchestrationError(f"親文書のMarkdownがありません: {parent_markdown}") from exc
    parent_markdown.write_text(
        original_markdown + "\n" + "\n".join(sections).rstrip() + "\n", encoding="utf-8"
    )


def _extract_xlsx_checklist_items(
    source: Path, original_path: str, header_row: int = 1
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise OrchestrationError("openpyxlをimportできません。preflightを再実行してください。") from exc
    try:
        workbook = load_workbook(source, data_only=False, read_only=True)
    except Exception as exc:
        raise OrchestrationError(f"チェックリストを読み取れません: {source} ({exc})") from exc
    items: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            headers = {
                cell.column: str(cell.value).strip()
                for cell in worksheet[header_row]
                if cell.value is not None and str(cell.value).strip()
            }
            for row in worksheet.iter_rows(min_row=header_row + 1):
                parts: list[str] = []
                for cell in row:
                    if cell.value is None or not str(cell.value).strip():
                        continue
                    label = headers.get(cell.column, cell.column_letter)
                    parts.append(f"{label}/{cell.coordinate}={cell.value}")
                if parts:
                    items.append(
                        {
                            "checklist_file": original_path,
                            "sheet": worksheet.title,
                            "row": row[0].row,
                            "check_item": " / ".join(parts),
                        }
                    )
    finally:
        workbook.close()
    if not items:
        raise OrchestrationError(
            f"チェックリストに{header_row + 1}行目以降のチェック項目がありません: {original_path}"
        )
    return items


def _validate_manifest_paths(root: Path, run_id: str, manifest: dict[str, Any]) -> None:
    if manifest.get("root") != ".":
        raise OrchestrationError("manifestのrootは`.`である必要があります")
    if manifest.get("run_id") != run_id:
        raise OrchestrationError("manifestのrun_idが指定値と一致しません")
    checklist_paths: set[str] = set()
    for role, input_prefix in ROLE_DIRECTORIES.items():
        entries = manifest.get(role)
        if not isinstance(entries, list):
            raise OrchestrationError(f"manifestの{role}は配列である必要があります")
        for index, entry in enumerate(entries):
            context = f"{role}[{index}]"
            if not isinstance(entry, dict):
                raise OrchestrationError(f"manifestの{context}が不正です")
            unknown_path_keys = {
                key
                for key in entry
                if key.endswith("_path")
                and key not in {"original_path", "intermediate_path"}
            }
            if unknown_path_keys:
                raise OrchestrationError(
                    f"manifestの{context}に未対応のパス項目があります: "
                    + ", ".join(sorted(unknown_path_keys))
                )
            original = _manifest_relative_path(
                root, entry.get("path"), input_prefix, f"{context}.path"
            )
            original_path = _manifest_relative_path(
                root,
                entry.get("original_path"),
                input_prefix,
                f"{context}.original_path",
            )
            if original != original_path:
                raise OrchestrationError(
                    f"manifestの{context}.pathとoriginal_pathが一致しません"
                )
            if role == "checklists":
                checklist_paths.add(original)
            _manifest_relative_path(
                root,
                entry.get("markdown"),
                f"work/markdown/{run_id}/{role}",
                f"{context}.markdown",
            )
            intermediate = entry.get("intermediate_path")
            if intermediate is not None:
                _manifest_relative_path(
                    root,
                    intermediate,
                    f"work/converted-office/{run_id}/{role}",
                    f"{context}.intermediate_path",
                )
    checklist_items = manifest.get("checklist_items")
    if not isinstance(checklist_items, list):
        raise OrchestrationError("manifestのchecklist_itemsは配列である必要があります")
    for index, item in enumerate(checklist_items):
        if not isinstance(item, dict):
            raise OrchestrationError(f"manifestのchecklist_items[{index}]が不正です")
        checklist_file = _manifest_relative_path(
            root,
            item.get("checklist_file"),
            "input/checklists",
            f"checklist_items[{index}].checklist_file",
        )
        if checklist_file not in checklist_paths:
            raise OrchestrationError(
                f"checklist_items[{index}].checklist_fileがchecklistsにありません"
            )


def _expected_results_path(
    root: Path, run_id: str, results_path: Path, *, require_relative: bool = False
) -> Path:
    supplied = Path(results_path)
    if ".." in supplied.parts or (require_relative and supplied.is_absolute()):
        raise OrchestrationError(
            "resultsはwork/review-runs/<run-id>/results.jsonのroot相対パスで指定してください"
        )
    candidate = _validated_repo_path(
        root, supplied if supplied.is_absolute() else root / supplied, "results"
    )
    expected = root / "work" / "review-runs" / run_id / "results.json"
    _validated_repo_path(root, expected, "results期待位置")
    if candidate != expected:
        raise OrchestrationError(
            "resultsは当該runのwork/review-runs/<run-id>/results.jsonを指定してください"
        )
    return candidate


def prepare_repository(root: Path, run_id: str, runner: Runner = _run_command) -> Path:
    root = root.resolve()
    _validate_run_id(run_id)
    _reject_existing_run_artifacts(root, run_id)
    preflight = _skills_root() / "review-documents-orchestrator" / "scripts" / "preflight.py"
    runner([sys.executable, str(preflight), "--root", str(root)], root)
    inventory = inventory_inputs(root)

    owned_paths, ownership_token = _create_owned_work_directories(root, run_id)
    try:
        return _prepare_repository_contents(root, run_id, inventory, runner)
    except Exception:
        for path in reversed(owned_paths):
            _cleanup_owned_directory(path, ownership_token)
        raise


def _prepare_repository_contents(
    root: Path, run_id: str, inventory: dict[str, list[Path]], runner: Runner
) -> Path:

    run_directory = root / "work" / "review-runs" / run_id
    manifest: dict[str, Any] = {
        "run_id": run_id,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "root": ".",
        "checklists": [],
        "references": [],
        "targets": [],
        "checklist_items": [],
    }

    legacy_script = _skills_root() / "convert-legacy-office" / "scripts" / "convert_legacy_office.py"
    for role, originals in inventory.items():
        for original in originals:
            original_relative = _relative(original, root)
            source = original
            intermediate_relative: str | None = None
            if original.suffix.lower() in LEGACY_FORMATS:
                source = _validated_repo_path(
                    root,
                    _legacy_destination(root, run_id, role, original),
                    "旧Office変換先",
                )
                source.parent.mkdir(parents=True, exist_ok=True)
                runner(
                    [
                        sys.executable,
                        str(legacy_script),
                        "convert",
                        str(original),
                        str(source),
                    ],
                    root,
                )
                intermediate_relative = _relative(source, root)
            markdown = _validated_repo_path(
                root,
                _markdown_destination(root, run_id, role, original),
                "Markdown生成先",
            )
            _validated_repo_path(
                root,
                _images_destination(root, run_id, role, original),
                "抽出画像生成先",
            )
            markdown.parent.mkdir(parents=True, exist_ok=True)
            runner(_format_command(root, run_id, role, source, original, markdown), root)
            if not markdown.is_file():
                raise OrchestrationError(f"Markdownが作られませんでした: {_relative(markdown, root)}")
            if intermediate_relative is not None:
                _rewrite_legacy_frontmatter(
                    markdown,
                    original_relative,
                    intermediate_relative,
                    original.suffix.lower(),
                    source.suffix.lower(),
                )
            if source.suffix.lower() in {".xlsx", ".docx", ".pptx"}:
                _ocr_embedded_images(
                    root,
                    run_id,
                    role,
                    original,
                    markdown,
                    _images_destination(root, run_id, role, original),
                    runner,
                )
            entry: dict[str, Any] = {
                "path": original_relative,
                "original_path": original_relative,
                "intermediate_path": intermediate_relative,
                "format": source.suffix.lower(),
                "markdown": _relative(markdown, root),
            }
            if role == "checklists":
                entry["header_row"] = 1
                manifest["checklist_items"].extend(
                    _extract_xlsx_checklist_items(source, original_relative, header_row=1)
                )
            manifest[role].append(entry)

    _validate_manifest_paths(root, run_id, manifest)
    manifest_path = _validated_repo_path(
        root, run_directory / "manifest.json", "manifest生成先"
    )
    _write_json(root, manifest_path, manifest)
    review_script = (
        _skills_root()
        / "review-markdown-documents"
        / "scripts"
        / "review_markdown_documents.py"
    )
    bundle_path = _validated_repo_path(
        root, run_directory / "review_bundle.json", "レビューbundle生成先"
    )
    runner(
        [
            sys.executable,
            str(review_script),
            "prepare",
            "--manifest",
            str(manifest_path),
            "--output",
            str(bundle_path),
        ],
        root,
    )
    if not bundle_path.is_file():
        raise OrchestrationError(f"レビューbundleが作られませんでした: {_relative(bundle_path, root)}")
    return manifest_path


def _unique_output_name(
    original_path: str, suffix: str, used_casefold_names: set[str]
) -> str:
    original = Path(original_path)
    name = original.with_suffix(suffix).name
    key = name.casefold()
    if key not in used_casefold_names:
        used_casefold_names.add(key)
        return name
    counter = 2
    while True:
        name = f"{original.stem}__{counter}{suffix}"
        key = name.casefold()
        if key not in used_casefold_names:
            break
        counter += 1
    used_casefold_names.add(key)
    return name


def _contract_entries(data: dict[str, Any], collection: str) -> dict[str, dict[str, Any]]:
    raw = data.get(collection)
    if not isinstance(raw, list):
        raise OrchestrationError(f"{collection}は配列である必要があります")
    entries: dict[str, dict[str, Any]] = {}
    for index, entry in enumerate(raw):
        if not isinstance(entry, dict) or not isinstance(entry.get("path"), str):
            raise OrchestrationError(f"{collection}[{index}]のpathが不正です")
        path = entry["path"]
        if path in entries:
            raise OrchestrationError(f"{collection}のpathが重複しています")
        entries[path] = entry
    return entries


def _checklist_item_keys(data: dict[str, Any]) -> dict[tuple[str, str, int], str]:
    raw = data.get("checklist_items")
    if not isinstance(raw, list):
        raise OrchestrationError("checklist_itemsは配列である必要があります")
    keys: dict[tuple[str, str, int], str] = {}
    for index, item in enumerate(raw):
        if not isinstance(item, dict):
            raise OrchestrationError(f"checklist_items[{index}]が不正です")
        key = (item.get("checklist_file"), item.get("sheet"), item.get("row"))
        if not isinstance(key[0], str) or not isinstance(key[1], str) or not isinstance(key[2], int):
            raise OrchestrationError(f"checklist_items[{index}]のキーが不正です")
        check_item = item.get("check_item")
        if not isinstance(check_item, str) or not check_item:
            raise OrchestrationError(f"checklist_items[{index}]のcheck_itemが不正です")
        if key in keys:
            raise OrchestrationError(f"checklist_itemsのキーが重複しています: {key}")
        keys[key] = check_item
    return keys


def _verify_results_match_manifest(manifest: dict[str, Any], results: dict[str, Any]) -> None:
    for collection in ("checklists", "references", "targets"):
        manifest_entries = _contract_entries(manifest, collection)
        result_entries = _contract_entries(results, collection)
        if set(manifest_entries) != set(result_entries):
            raise OrchestrationError(
                f"resultsの{collection}がmanifestと一致しません。別runの結果を混在させないでください。"
            )
        for path, manifest_entry in manifest_entries.items():
            result_entry = result_entries[path]
            if manifest_entry.get("markdown") != result_entry.get("markdown"):
                raise OrchestrationError(
                    f"resultsの{collection}にあるMarkdownパスがmanifestと一致しません: {path}"
                )
            if collection == "checklists" and manifest_entry.get("header_row", 1) != result_entry.get(
                "header_row", 1
            ):
                raise OrchestrationError(
                    f"resultsのチェックリストheader_rowがmanifestと一致しません: {path}"
                )
    if _checklist_item_keys(manifest) != _checklist_item_keys(results):
        raise OrchestrationError(
            "resultsのchecklist_itemsがmanifestと一致しません。別runの結果を混在させないでください。"
        )


def finalize_repository(
    root: Path, run_id: str, results_path: Path, runner: Runner = _run_command
) -> Path:
    root = root.resolve()
    _validate_run_id(run_id)
    manifest_path = _validated_repo_path(
        root,
        root / "work" / "review-runs" / run_id / "manifest.json",
        "manifest",
    )
    manifest = _read_json(manifest_path)
    _validate_manifest_paths(root, run_id, manifest)
    results_path = _expected_results_path(root, run_id, results_path)
    results = _read_json(results_path)
    if results.get("run_id") != run_id:
        raise OrchestrationError("resultsのrun_idが指定値と一致しません")

    review_script = (
        _skills_root()
        / "review-markdown-documents"
        / "scripts"
        / "review_markdown_documents.py"
    )
    runner(
        [sys.executable, str(review_script), "validate", "--results", str(results_path)],
        root,
    )
    _verify_results_match_manifest(manifest, results)
    output_directory = _validated_repo_path(
        root, root / "output" / "reviews" / run_id, "レビュー結果生成先"
    )
    if output_directory.exists() or _is_link_like(output_directory):
        raise OrchestrationError(
            f"同じrun-idの結果フォルダが既にあります: {_relative(output_directory, root)}"
        )
    output_directory.parent.mkdir(parents=True, exist_ok=True)
    staging_directory = _validated_repo_path(
        root,
        root / "work" / "review-runs" / run_id / "finalize-staging",
        "確定処理staging生成先",
    )
    if staging_directory.exists() or _is_link_like(staging_directory):
        raise OrchestrationError(
            f"前回失敗した確定処理の作業フォルダがあります: {_relative(staging_directory, root)}"
        )
    staging_directory.mkdir(parents=True)
    ownership_token = uuid.uuid4().hex
    try:
        (staging_directory / OWNER_MARKER).write_text(
            ownership_token, encoding="utf-8"
        )
    except Exception:
        staging_directory.rmdir()
        raise
    try:
        return _finalize_repository_contents(
            root,
            manifest,
            results_path,
            review_script,
            output_directory,
            staging_directory,
            runner,
        )
    except Exception:
        _cleanup_owned_directory(staging_directory, ownership_token)
        raise


def _finalize_repository_contents(
    root: Path,
    manifest: dict[str, Any],
    results_path: Path,
    review_script: Path,
    output_directory: Path,
    staging_directory: Path,
    runner: Runner,
) -> Path:

    summary_path = _validated_repo_path(
        root, staging_directory / "summary.md", "summary生成先"
    )
    runner(
        [
            sys.executable,
            str(review_script),
            "summary",
            "--results",
            str(results_path),
            "--output",
            str(summary_path),
        ],
        root,
    )
    if not summary_path.is_file():
        raise OrchestrationError("summary.mdが作られませんでした")
    xlsx_script = _skills_root() / "xlsx-document" / "scripts" / "xlsx_document.py"
    used_casefold_names: set[str] = set()
    for checklist in manifest.get("checklists", []):
        if not isinstance(checklist, dict):
            raise OrchestrationError("manifestのchecklists要素が不正です")
        original_relative = checklist.get("original_path") or checklist.get("path")
        if not isinstance(original_relative, str):
            raise OrchestrationError("チェックリストのoriginal_pathがありません")
        intermediate = checklist.get("intermediate_path")
        source = root / intermediate if isinstance(intermediate, str) and intermediate else root / original_relative
        output_name = _unique_output_name(
            original_relative, ".xlsx", used_casefold_names
        )
        destination = _validated_repo_path(
            root, staging_directory / output_name, "結果チェックリスト生成先"
        )
        runner(
            [
                sys.executable,
                str(xlsx_script),
                "write-review",
                str(source),
                str(destination),
                "--results",
                str(results_path),
                "--checklist-path",
                original_relative,
            ],
            root,
        )
        if not destination.is_file():
            raise OrchestrationError(f"結果チェックリストが作られませんでした: {destination.name}")
    (staging_directory / OWNER_MARKER).unlink()
    try:
        staging_directory.rename(output_directory)
    except Exception:
        if staging_directory.is_dir() and not _is_link_like(staging_directory):
            shutil.rmtree(staging_directory)
        raise
    return output_directory


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="設計書レビュー処理の統合実行")
    subparsers = parser.add_subparsers(dest="command", required=True)
    prepare = subparsers.add_parser("prepare", help="入力をMarkdown化してレビューbundleを作る")
    prepare.add_argument("--root", type=Path, required=True)
    prepare.add_argument("--run-id")
    finalize = subparsers.add_parser("finalize", help="検証済み結果から成果物を作る")
    finalize.add_argument("--root", type=Path, required=True)
    finalize.add_argument("--run-id", required=True)
    finalize.add_argument("--results", type=Path, required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        if args.command == "prepare":
            run_id = args.run_id or datetime.now().astimezone().strftime("%Y%m%d%H%M")
            manifest_path = prepare_repository(args.root, run_id)
            print(
                "レビュー材料を準備しました: "
                f"{_relative(manifest_path, args.root.resolve())}"
            )
            print(f"run-id: {run_id}")
        else:
            _expected_results_path(
                args.root.resolve(),
                args.run_id,
                args.results,
                require_relative=True,
            )
            output_directory = finalize_repository(
                args.root, args.run_id, args.results
            )
            print(
                "レビュー結果を作成しました: "
                f"{_relative(output_directory, args.root.resolve())}"
            )
        return 0
    except OrchestrationError as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
