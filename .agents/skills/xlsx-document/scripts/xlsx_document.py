#!/usr/bin/env python3
"""XLSXを位置情報付きMarkdownへ変換し、安全な限定編集を行う。"""

from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
import json
import os
from pathlib import Path
import re
import sys
from typing import Any


ALLOWED_RESULTS = {"適合", "不適合", "対象外", "要確認"}


def _require_runtime() -> None:
    if sys.version_info < (3, 12):
        raise RuntimeError(
            f"Python 3.12以上が必要です。現在: {sys.version.split()[0]}"
        )


def _openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "必要なライブラリ openpyxl を読み込めません。READMEのインストール手順を確認してください。"
        ) from exc
    return openpyxl


def _ensure_distinct(input_path: Path, output_path: Path) -> None:
    if input_path.resolve() == output_path.resolve():
        raise ValueError("入力ファイルは上書きできません。別の出力パスを指定してください。")
    _validate_new_output_path(output_path, "出力先")


def _reject_symlink_components(path: Path, label: str) -> None:
    """出力パス自体と既存祖先のsymlinkを拒否する。"""
    absolute = path if path.is_absolute() else Path.cwd() / path
    for candidate in (absolute, *absolute.parents):
        if _is_link_like(candidate):
            raise ValueError(
                f"{label}にシンボリックリンクを使用できません: {candidate}"
            )


def _is_link_like(path: Path) -> bool:
    if path.is_symlink():
        return True
    is_junction = getattr(path, "is_junction", None)
    if is_junction is None:
        return False
    try:
        return bool(is_junction())
    except OSError:
        return False


def _validate_new_output_path(path: Path, label: str) -> None:
    _reject_symlink_components(path, label)
    if os.path.lexists(path):
        raise FileExistsError(
            f"{label}は既に存在します。上書きせず、何も書き込みません: {path}"
        )


def _validate_output_directory(path: Path, label: str) -> None:
    _reject_symlink_components(path, label)
    if os.path.lexists(path) and not path.is_dir():
        raise NotADirectoryError(f"{label}はディレクトリではありません: {path}")


def _validate_input(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"XLSXファイルを指定してください: {path}")


def _repo_relative(path: Path, repo_root: Path) -> str:
    try:
        return path.resolve().relative_to(repo_root.resolve()).as_posix()
    except ValueError as exc:
        raise ValueError(
            f"入力ファイルはリポジトリルート配下に置いてください: {path}"
        ) from exc


def _yaml_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def _md(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.replace("|", "\\|").replace("\n", "<br>")


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _merged_ranges_by_cell(worksheet: Any) -> dict[str, str]:
    result: dict[str, str] = {}
    for merged in worksheet.merged_cells.ranges:
        merged_text = str(merged)
        for row in worksheet[merged_text]:
            for cell in row:
                result[cell.coordinate] = merged_text
    return result


def _plan_images(
    workbook: Any, input_path: Path, images_dir: Path | None
) -> dict[str, list[dict[str, Any]]]:
    """全シートの画像出力を先に確定し、衝突時は書込み前に停止する。"""
    plans: dict[str, list[dict[str, Any]]] = {}
    planned_paths: set[str] = set()
    conflicts: list[Path] = []
    for worksheet in workbook.worksheets:
        sheet_plans: list[dict[str, Any]] = []
        safe_sheet = re.sub(r"[^\w.-]+", "_", worksheet.title, flags=re.UNICODE).strip("._") or "sheet"
        for image_index, image in enumerate(getattr(worksheet, "_images", []), start=1):
            anchor = getattr(image, "anchor", None)
            location = "不明"
            if hasattr(anchor, "_from"):
                location = f"行{anchor._from.row + 1}・列{anchor._from.col + 1}"
            plan: dict[str, Any] = {
                "index": image_index,
                "location": location,
                "path": None,
                "data": None,
            }
            if images_dir is not None:
                image_format = str(getattr(image, "format", "") or "png").lower()
                image_format = "jpg" if image_format == "jpeg" else image_format
                if not re.fullmatch(r"[a-z0-9]+", image_format):
                    image_format = "bin"
                image_path = images_dir / (
                    f"{input_path.stem}-sheet-{safe_sheet}-image-{image_index:04d}.{image_format}"
                )
                _validate_new_output_path(image_path, "抽出画像の出力先")
                path_key = str(image_path.resolve()).casefold()
                if path_key in planned_paths:
                    conflicts.append(image_path)
                planned_paths.add(path_key)
                plan["path"] = image_path
                plan["data"] = image._data()
            sheet_plans.append(plan)
        plans[worksheet.title] = sheet_plans
    if conflicts:
        joined = "\n- ".join(str(path) for path in conflicts)
        raise FileExistsError(f"抽出画像の出力先が重複または既存です。何も書き込みません:\n- {joined}")
    return plans


def to_markdown(
    input_path: Path,
    output_path: Path,
    role: str,
    repo_root: Path,
    images_dir: Path | None,
) -> None:
    openpyxl = _openpyxl()
    _validate_input(input_path)
    _ensure_distinct(input_path, output_path)
    source_path = _repo_relative(input_path, repo_root)
    if images_dir is not None:
        _validate_output_directory(images_dir, "画像出力ディレクトリ")

    workbook = openpyxl.load_workbook(input_path, data_only=False, read_only=False)
    values_workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=False)
    image_plans = _plan_images(workbook, input_path, images_dir)
    lines = [
        "---",
        f"source_path: {_yaml_string(source_path)}",
        f"source_name: {_yaml_string(input_path.name)}",
        'source_format: "xlsx"',
        f"document_role: {_yaml_string(role)}",
        f"converted_at: {_yaml_string(datetime.now().astimezone().isoformat(timespec='seconds'))}",
        'converter_skill: "xlsx-document"',
        "---",
        "",
        f"# {_md(input_path.name)}",
        "",
    ]

    for worksheet in workbook.worksheets:
        values_sheet = values_workbook[worksheet.title]
        merged_by_cell = _merged_ranges_by_cell(worksheet)
        merged_ranges = [str(item) for item in worksheet.merged_cells.ranges]
        lines.extend(
            [
                f"## シート: {_md(worksheet.title)}",
                "",
                f"- 位置: `シート {worksheet.title}`",
                f"- 使用範囲: `{worksheet.calculate_dimension()}`",
                f"- 結合セル: {_md(', '.join(merged_ranges) if merged_ranges else 'なし')}",
                "",
                "| セル位置 | 値 | 種別 | 数式 | コメント | 結合範囲 |",
                "|---|---|---|---|---|---|",
            ]
        )

        found = False
        for row in worksheet.iter_rows():
            for cell in row:
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else ""
                cached_value = values_sheet[cell.coordinate].value if formula else cell.value
                comment = ""
                if cell.comment is not None:
                    author = f" ({cell.comment.author})" if cell.comment.author else ""
                    comment = f"{cell.comment.text}{author}"
                merged = merged_by_cell.get(cell.coordinate, "")
                if cell.value is None and not comment and not merged:
                    continue
                found = True
                value_type = "数式" if formula else (cell.data_type or type(cell.value).__name__)
                lines.append(
                    "| "
                    + " | ".join(
                        [
                            f"`{cell.coordinate}`",
                            _md(_display_value(cached_value)),
                            _md(value_type),
                            _md(formula),
                            _md(comment),
                            _md(merged),
                        ]
                    )
                    + " |"
                )
        if not found:
            lines.append("| - | - | - | - | - | - |")

        sheet_image_plans = image_plans[worksheet.title]
        if sheet_image_plans:
            lines.extend(["", "### 埋め込み画像", ""])
            for image_plan in sheet_image_plans:
                image_index = image_plan["index"]
                location = image_plan["location"]
                image_reference = "抽出先未指定"
                image_path = image_plan["path"]
                if image_path is not None:
                    relative_link = Path(os.path.relpath(image_path, output_path.parent)).as_posix()
                    image_reference = f"![シート画像 {image_index}]({relative_link})"
                lines.append(
                    f"- 画像 {image_index}: シート `{worksheet.title}`、アンカー `{location}`、{image_reference}"
                )
        lines.append("")

    for sheet_plans in image_plans.values():
        for image_plan in sheet_plans:
            image_path = image_plan["path"]
            if image_path is None:
                continue
            image_path.parent.mkdir(parents=True, exist_ok=True)
            with image_path.open("xb") as stream:
                stream.write(image_plan["data"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("x", encoding="utf-8", newline="\n") as stream:
        stream.write("\n".join(lines).rstrip() + "\n")
    workbook.close()
    values_workbook.close()


def _load_operations(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"操作JSONが見つかりません: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"操作JSONを読み取れません: {path}: {exc}") from exc
    operations = payload.get("operations") if isinstance(payload, dict) else payload
    if not isinstance(operations, list):
        raise ValueError("操作JSONは配列、または operations 配列を持つオブジェクトにしてください。")
    if not all(isinstance(item, dict) for item in operations):
        raise ValueError("各操作はJSONオブジェクトにしてください。")
    return operations


def edit(input_path: Path, output_path: Path, operations_path: Path) -> None:
    openpyxl = _openpyxl()
    _validate_input(input_path)
    _ensure_distinct(input_path, output_path)
    workbook = openpyxl.load_workbook(input_path, data_only=False, read_only=False)

    for index, operation in enumerate(_load_operations(operations_path), start=1):
        operation_name = operation.get("op")
        try:
            if operation_name == "set_cell":
                sheet = operation["sheet"]
                coordinate = operation["cell"]
                if sheet not in workbook.sheetnames:
                    raise ValueError(f"シートが見つかりません: {sheet}")
                workbook[sheet][coordinate] = operation.get("value")
            elif operation_name == "rename_sheet":
                sheet = operation["sheet"]
                new_name = operation["new_name"]
                if sheet not in workbook.sheetnames:
                    raise ValueError(f"シートが見つかりません: {sheet}")
                if new_name in workbook.sheetnames and new_name != sheet:
                    raise ValueError(f"同名のシートが存在します: {new_name}")
                workbook[sheet].title = new_name
            elif operation_name == "add_sheet":
                name = operation["name"]
                if name in workbook.sheetnames:
                    raise ValueError(f"同名のシートが存在します: {name}")
                position = operation.get("index")
                if position is not None and (not isinstance(position, int) or position < 0):
                    raise ValueError("add_sheet の index は0以上の整数にしてください。")
                workbook.create_sheet(title=name, index=position)
            else:
                raise ValueError(f"許可されていない操作です: {operation_name}")
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"操作 {index} が不正です: {exc}") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def _normal_path(value: Any) -> str:
    return Path(str(value).replace("\\", "/")).as_posix().lstrip("./")


def _load_results(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(f"結果JSONが見つかりません: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"結果JSONを読み取れません: {path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError("結果JSONのルートはオブジェクトにしてください。")
    for key in ("checklists", "targets", "items"):
        if not isinstance(payload.get(key), list):
            raise ValueError(f"結果JSONに {key} 配列が必要です。")
    return payload


def _copy_cell_style(source: Any, destination: Any) -> None:
    destination.font = copy(source.font)
    destination.fill = copy(source.fill)
    destination.border = copy(source.border)
    destination.alignment = copy(source.alignment)
    destination.number_format = source.number_format
    destination.protection = copy(source.protection)


def _last_used_column(worksheet: Any) -> int:
    last = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None or cell.comment is not None:
                last = max(last, cell.column)
    for merged in worksheet.merged_cells.ranges:
        last = max(last, merged.max_col)
    return max(last, 1)


def _result_comment(item: dict[str, Any]) -> str:
    comment = str(item.get("comment", "")).strip()
    evidence = item.get("evidence", [])
    if not comment:
        raise ValueError("レビューコメントは空にできません。")
    if not isinstance(evidence, list):
        raise ValueError("evidence は配列にしてください。")
    evidence_text = "; ".join(str(value).strip() for value in evidence if str(value).strip())
    return f"{comment}\n根拠: {evidence_text}" if evidence_text else comment


def write_review(
    input_path: Path,
    output_path: Path,
    results_path: Path,
    checklist_path: str,
) -> None:
    openpyxl = _openpyxl()
    _validate_input(input_path)
    _ensure_distinct(input_path, output_path)
    payload = _load_results(results_path)
    normalized_checklist = _normal_path(checklist_path)

    checklist_entries = [
        entry
        for entry in payload["checklists"]
        if isinstance(entry, dict) and _normal_path(entry.get("path", "")) == normalized_checklist
    ]
    if len(checklist_entries) != 1:
        raise ValueError(
            f"checklists に対象パスを1件だけ指定してください: {normalized_checklist}"
        )
    header_row = checklist_entries[0].get("header_row", 1)
    if not isinstance(header_row, int) or header_row < 1:
        raise ValueError("header_row は1以上の整数にしてください。")

    targets: list[str] = []
    for target in payload["targets"]:
        if not isinstance(target, dict) or not str(target.get("path", "")).strip():
            raise ValueError("targets の各要素には path が必要です。")
        target_path = _normal_path(target["path"])
        if target_path in targets:
            raise ValueError(f"targets の path が重複しています: {target_path}")
        targets.append(target_path)

    indexed: dict[tuple[str, int, str], dict[str, Any]] = {}
    for item_index, item in enumerate(payload["items"], start=1):
        if not isinstance(item, dict):
            raise ValueError(f"items[{item_index - 1}] はオブジェクトにしてください。")
        if _normal_path(item.get("checklist_file", "")) != normalized_checklist:
            continue
        sheet = str(item.get("sheet", ""))
        row = item.get("row")
        target = _normal_path(item.get("target_file", ""))
        result = item.get("result")
        if not sheet or not isinstance(row, int) or row < 1:
            raise ValueError(f"items[{item_index - 1}] の sheet または row が不正です。")
        if target not in targets:
            raise ValueError(f"items[{item_index - 1}] の target_file が targets にありません: {target}")
        if result not in ALLOWED_RESULTS:
            raise ValueError(
                f"items[{item_index - 1}] の result は適合/不適合/対象外/要確認から選んでください。"
            )
        _result_comment(item)
        key = (sheet, row, target)
        if key in indexed:
            raise ValueError(f"同じレビュー結果が重複しています: {sheet}!{row}, {target}")
        indexed[key] = item

    workbook = openpyxl.load_workbook(input_path, data_only=False, read_only=False)
    unknown_sheets = sorted({sheet for sheet, _, _ in indexed if sheet not in workbook.sheetnames})
    if unknown_sheets:
        workbook.close()
        raise ValueError(f"チェックリストに存在しないシートです: {', '.join(unknown_sheets)}")

    base_columns = {worksheet.title: _last_used_column(worksheet) for worksheet in workbook.worksheets}
    for worksheet in workbook.worksheets:
        base_column = base_columns[worksheet.title]
        for target_index, target in enumerate(targets):
            start_column = base_column + target_index * 3 + 1
            headers = ("レビュー対象ファイル名", "レビュー結果", "レビューコメント")
            for offset, header in enumerate(headers):
                destination = worksheet.cell(row=header_row, column=start_column + offset, value=header)
                _copy_cell_style(worksheet.cell(row=header_row, column=base_column), destination)
                source_letter = openpyxl.utils.get_column_letter(base_column)
                destination_letter = openpyxl.utils.get_column_letter(start_column + offset)
                source_width = worksheet.column_dimensions[source_letter].width
                worksheet.column_dimensions[destination_letter].width = source_width
                if offset == 2:
                    worksheet.column_dimensions[destination_letter].width = max(source_width or 0, 40)

            for (sheet, row, item_target), item in indexed.items():
                if sheet != worksheet.title or item_target != target:
                    continue
                values = (Path(target).name, item["result"], _result_comment(item))
                for offset, value in enumerate(values):
                    destination = worksheet.cell(row=row, column=start_column + offset, value=value)
                    _copy_cell_style(worksheet.cell(row=row, column=base_column), destination)
                    alignment = copy(destination.alignment)
                    alignment.wrap_text = True
                    destination.alignment = alignment

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    markdown_parser = subparsers.add_parser("to-markdown", help="XLSXをMarkdownへ変換する")
    markdown_parser.add_argument("input", type=Path)
    markdown_parser.add_argument("output", type=Path)
    markdown_parser.add_argument("--role", required=True, choices=("checklist", "reference", "target"))
    markdown_parser.add_argument("--repo-root", required=True, type=Path)
    markdown_parser.add_argument("--images-dir", type=Path, help="埋め込み画像の抽出先")

    edit_parser = subparsers.add_parser("edit", help="許可した操作だけでXLSXを編集する")
    edit_parser.add_argument("input", type=Path)
    edit_parser.add_argument("output", type=Path)
    edit_parser.add_argument("--operations", required=True, type=Path)

    review_parser = subparsers.add_parser("write-review", help="チェックリストのコピーへレビュー結果を記入する")
    review_parser.add_argument("input", type=Path)
    review_parser.add_argument("output", type=Path)
    review_parser.add_argument("--results", required=True, type=Path)
    review_parser.add_argument("--checklist-path", required=True)
    return parser


def main() -> int:
    try:
        _require_runtime()
        arguments = build_parser().parse_args()
        if arguments.command == "to-markdown":
            to_markdown(
                arguments.input,
                arguments.output,
                arguments.role,
                arguments.repo_root,
                arguments.images_dir,
            )
        elif arguments.command == "edit":
            edit(arguments.input, arguments.output, arguments.operations)
        else:
            write_review(
                arguments.input,
                arguments.output,
                arguments.results,
                arguments.checklist_path,
            )
        return 0
    except Exception as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
