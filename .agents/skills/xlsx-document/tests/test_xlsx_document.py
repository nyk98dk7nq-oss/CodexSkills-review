from __future__ import annotations

import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

import openpyxl
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.styles import Font, PatternFill
from PIL import Image


SCRIPT = Path(__file__).resolve().parents[1] / "scripts" / "xlsx_document.py"


class XlsxDocumentCliTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        (self.root / "input" / "checklists").mkdir(parents=True)
        self.source = self.root / "input" / "checklists" / "checklist.xlsx"
        image_path = self.root / "input" / "checklists" / "sample.png"
        Image.new("RGB", (18, 12), "red").save(image_path)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "確認"
        sheet["A1"] = "項目"
        sheet["A1"].font = Font(bold=True)
        sheet["A1"].fill = PatternFill("solid", fgColor="FFFF00")
        sheet["A2"] = "合計が正しい"
        sheet["B2"] = 2
        sheet["C2"] = 3
        sheet["D2"] = "=B2+C2"
        sheet["A2"].comment = Comment("確認条件", "reviewer")
        sheet.merge_cells("A3:B3")
        sheet["A3"] = "結合"
        sheet.add_image(SpreadsheetImage(str(image_path)), "F2")
        sheet.add_image(SpreadsheetImage(str(image_path)), "G2")
        workbook.save(self.source)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def run_cli(self, *arguments: object) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(SCRIPT), *(str(value) for value in arguments)],
            text=True,
            capture_output=True,
            check=False,
        )

    def make_symlink(
        self, link: Path, target: Path, *, target_is_directory: bool = False
    ) -> None:
        try:
            link.symlink_to(target, target_is_directory=target_is_directory)
        except (NotImplementedError, OSError) as exc:
            self.skipTest(f"シンボリックリンクを作成できません: {exc}")

    def test_to_markdown_preserves_locations_formula_comment_and_merge(self) -> None:
        output = self.root / "work" / "checklist.md"
        images = self.root / "work" / "images"
        result = self.run_cli(
            "to-markdown",
            self.source,
            output,
            "--role",
            "checklist",
            "--repo-root",
            self.root,
            "--images-dir",
            images,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        text = output.read_text(encoding="utf-8")
        self.assertIn('source_path: "input/checklists/checklist.xlsx"', text)
        self.assertIn('source_name: "checklist.xlsx"', text)
        self.assertIn('document_role: "checklist"', text)
        self.assertIn('converter_skill: "xlsx-document"', text)
        self.assertIn("## シート: 確認", text)
        self.assertIn("`D2`", text)
        self.assertIn("=B2+C2", text)
        self.assertIn("確認条件 (reviewer)", text)
        self.assertIn("A3:B3", text)
        self.assertIn("アンカー `行2・列6`", text)
        self.assertIn("![シート画像 1]", text)
        self.assertEqual(len(list(images.glob("checklist-sheet-確認-image-*.png"))), 2)

    def test_image_conflict_stops_before_writing_any_output(self) -> None:
        output = self.root / "work" / "checklist.md"
        images = self.root / "input" / "checklists"
        conflict = images / "checklist-sheet-確認-image-0002.png"
        sentinel = "利用者の既存画像".encode("utf-8")
        conflict.write_bytes(sentinel)
        result = self.run_cli(
            "to-markdown",
            self.source,
            output,
            "--role",
            "checklist",
            "--repo-root",
            self.root,
            "--images-dir",
            images,
        )
        self.assertNotEqual(result.returncode, 0)
        self.assertEqual(conflict.read_bytes(), sentinel)
        self.assertFalse((images / "checklist-sheet-確認-image-0001.png").exists())
        self.assertFalse(output.exists())

    def test_broken_planned_image_symlink_stops_all_output(self) -> None:
        output = self.root / "work" / "checklist.md"
        images = self.root / "work" / "images"
        images.mkdir(parents=True)
        broken = images / "checklist-sheet-確認-image-0002.png"
        self.make_symlink(broken, self.root / "outside" / "missing.png")

        result = self.run_cli(
            "to-markdown",
            self.source,
            output,
            "--role",
            "checklist",
            "--repo-root",
            self.root,
            "--images-dir",
            images,
        )

        self.assertNotEqual(result.returncode, 0)
        self.assertIn("シンボリックリンク", result.stderr)
        self.assertTrue(broken.is_symlink())
        self.assertFalse((images / "checklist-sheet-確認-image-0001.png").exists())
        self.assertFalse(output.exists())

    def test_symlink_primary_output_parent_and_images_dir_are_rejected(self) -> None:
        work = self.root / "work"
        work.mkdir()
        broken_output = work / "broken.md"
        broken_target = self.root / "outside" / "missing.md"
        self.make_symlink(broken_output, broken_target)
        result = self.run_cli(
            "to-markdown",
            self.source,
            broken_output,
            "--role",
            "checklist",
            "--repo-root",
            self.root,
        )
        self.assertNotEqual(result.returncode, 0)
        self.assertTrue(broken_output.is_symlink())
        self.assertFalse(broken_target.exists())

        outside = self.root / "outside"
        outside.mkdir()
        output_parent_link = self.root / "edited-link"
        self.make_symlink(output_parent_link, outside, target_is_directory=True)
        operations = self.root / "empty-operations.json"
        operations.write_text("[]", encoding="utf-8")
        edited = output_parent_link / "edited.xlsx"
        result = self.run_cli("edit", self.source, edited, "--operations", operations)
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("シンボリックリンク", result.stderr)
        self.assertFalse((outside / "edited.xlsx").exists())

        images_link = self.root / "images-link"
        self.make_symlink(images_link, outside, target_is_directory=True)
        markdown = work / "safe.md"
        result = self.run_cli(
            "to-markdown",
            self.source,
            markdown,
            "--role",
            "checklist",
            "--repo-root",
            self.root,
            "--images-dir",
            images_link,
        )
        self.assertNotEqual(result.returncode, 0)
        self.assertIn("シンボリックリンク", result.stderr)
        self.assertFalse(markdown.exists())
        self.assertEqual(list(outside.iterdir()), [])

    def test_all_commands_reject_existing_primary_output(self) -> None:
        markdown = self.root / "work" / "existing.md"
        markdown.parent.mkdir(parents=True)
        markdown.write_text("保持", encoding="utf-8")
        converted = self.run_cli(
            "to-markdown",
            self.source,
            markdown,
            "--role",
            "checklist",
            "--repo-root",
            self.root,
        )
        self.assertNotEqual(converted.returncode, 0)
        self.assertEqual(markdown.read_text(encoding="utf-8"), "保持")

        existing_book = self.root / "output" / "existing.xlsx"
        existing_book.parent.mkdir(parents=True)
        existing_book.write_bytes("保持".encode("utf-8"))
        operations = self.root / "empty-operations.json"
        operations.write_text("[]", encoding="utf-8")
        edited = self.run_cli("edit", self.source, existing_book, "--operations", operations)
        self.assertNotEqual(edited.returncode, 0)
        self.assertEqual(existing_book.read_bytes(), "保持".encode("utf-8"))

        reviewed = self.run_cli(
            "write-review",
            self.source,
            existing_book,
            "--results",
            self.root / "missing-results.json",
            "--checklist-path",
            "input/checklists/checklist.xlsx",
        )
        self.assertNotEqual(reviewed.returncode, 0)
        self.assertEqual(existing_book.read_bytes(), "保持".encode("utf-8"))

    def test_edit_applies_only_supported_operations(self) -> None:
        operations = self.root / "operations.json"
        operations.write_text(
            json.dumps(
                {
                    "operations": [
                        {"op": "set_cell", "sheet": "確認", "cell": "B2", "value": 9},
                        {"op": "rename_sheet", "sheet": "確認", "new_name": "レビュー"},
                        {"op": "add_sheet", "name": "履歴"},
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        output = self.root / "output" / "edited.xlsx"
        result = self.run_cli("edit", self.source, output, "--operations", operations)
        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(output, data_only=False)
        self.assertEqual(workbook["レビュー"]["B2"].value, 9)
        self.assertIn("履歴", workbook.sheetnames)
        workbook.close()
        original = openpyxl.load_workbook(self.source)
        self.assertEqual(original["確認"]["B2"].value, 2)
        original.close()

    def test_write_review_adds_three_columns_per_target(self) -> None:
        results = self.root / "results.json"
        results.write_text(
            json.dumps(
                {
                    "checklists": [
                        {"path": "input/checklists/checklist.xlsx", "header_row": 1}
                    ],
                    "targets": [
                        {"path": "input/targets/design.docx", "markdown": "work/design.md"}
                    ],
                    "items": [
                        {
                            "checklist_file": "input/checklists/checklist.xlsx",
                            "sheet": "確認",
                            "row": 2,
                            "check_item": "合計が正しい",
                            "target_file": "input/targets/design.docx",
                            "result": "適合",
                            "comment": "設計書の表2で合計値を確認した。",
                            "evidence": ["work/design.md:表2"],
                            "improvement": "",
                        }
                    ],
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        output = self.root / "output" / "reviewed.xlsx"
        result = self.run_cli(
            "write-review",
            self.source,
            output,
            "--results",
            results,
            "--checklist-path",
            "input/checklists/checklist.xlsx",
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook["確認"]
        self.assertEqual(sheet["E1"].value, "レビュー対象ファイル名")
        self.assertEqual(sheet["F1"].value, "レビュー結果")
        self.assertEqual(sheet["G1"].value, "レビューコメント")
        self.assertEqual(sheet["E2"].value, "design.docx")
        self.assertEqual(sheet["F2"].value, "適合")
        self.assertIn("work/design.md:表2", sheet["G2"].value)
        self.assertEqual(sheet["E1"].font.bold, sheet["D1"].font.bold)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
