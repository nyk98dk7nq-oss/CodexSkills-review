#!/usr/bin/env python3
"""Excelブックを安全に読み取り、Markdown変換用の決定的なJSONを生成する。"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import posixpath
import re
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ElementTree
from datetime import date, datetime, time, timedelta
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
from urllib.parse import unquote

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError as exc:  # pragma: no cover - 実行環境依存
    print(
        "エラー: inspect_excel.pyにはopenpyxl>=3.1,<3.2が必要です。",
        file=sys.stderr,
    )
    raise SystemExit(3) from exc

_OPENPYXL_VERSION = re.match(r"^(\d+)\.(\d+)", openpyxl.__version__)
if (
    _OPENPYXL_VERSION is None
    or tuple(map(int, _OPENPYXL_VERSION.groups())) != (3, 1)
):
    print(
        "エラー: openpyxl>=3.1,<3.2が必要です。"
        f"現在: {openpyxl.__version__}",
        file=sys.stderr,
    )
    raise SystemExit(3)


SCHEMA_VERSION = "1.1"
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
DEFAULT_MAX_FILE_MB = 25
DEFAULT_MAX_UNCOMPRESSED_MB = 250
DEFAULT_MAX_CELLS = 100_000
DEFAULT_MAX_TOTAL_CELLS = 250_000
MAX_ARCHIVE_PARTS = 20_000
MAX_SUSPICIOUS_COMPRESSION_RATIO = 1_000
MAX_CONTENT_TYPES_BYTES = 5 * 1024 * 1024
MAX_RELATIONSHIPS_BYTES = 10 * 1024 * 1024


class InspectionError(RuntimeError):
    """ユーザーが対処できる検査エラー。"""


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Excel（.xlsx/.xlsm）を解析し、Markdown変換用JSONを出力します。"
    )
    parser.add_argument("workbook", help="解析するExcelブック")
    parser.add_argument(
        "--output",
        "-o",
        help="JSONの出力先。省略時は標準出力へ出力します。",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="既存のJSON出力を上書きします。--output指定時だけ使用できます。",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        default=[],
        help="抽出するシート名。複数回指定できます。",
    )
    parser.add_argument(
        "--include-hidden-sheets",
        action="store_true",
        help="非表示・veryHiddenシートも抽出対象に含めます。",
    )
    parser.add_argument(
        "--include-hidden-rows-columns",
        action="store_true",
        help="対象シートの非表示行・列にあるセル内容も抽出します。",
    )
    parser.add_argument(
        "--no-styles",
        action="store_true",
        help="セルのスタイル情報を省略してJSONを小さくします。",
    )
    parser.add_argument(
        "--max-file-mb",
        type=int,
        default=DEFAULT_MAX_FILE_MB,
        help=f"入力ファイルの上限MiB（既定: {DEFAULT_MAX_FILE_MB}）。",
    )
    parser.add_argument(
        "--max-uncompressed-mb",
        type=int,
        default=DEFAULT_MAX_UNCOMPRESSED_MB,
        help=f"ZIP展開後の合計サイズ上限MiB（既定: {DEFAULT_MAX_UNCOMPRESSED_MB}）。",
    )
    parser.add_argument(
        "--max-cells",
        "--max-grid-cells",
        dest="max_cells",
        type=int,
        default=DEFAULT_MAX_CELLS,
        help=f"1シートから抽出する内容セル数上限（既定: {DEFAULT_MAX_CELLS}）。",
    )
    parser.add_argument(
        "--max-total-cells",
        "--max-total-grid-cells",
        dest="max_total_cells",
        type=int,
        default=DEFAULT_MAX_TOTAL_CELLS,
        help=f"対象シート合計の内容セル数上限（既定: {DEFAULT_MAX_TOTAL_CELLS}）。",
    )
    return parser.parse_args(argv)


def path_entry_exists(path: Path) -> bool:
    """壊れたシンボリックリンクを含め、パスのディレクトリエントリを調べる。"""
    return os.path.lexists(os.fspath(path))


def resolved_path(path: Path, label: str) -> Path:
    try:
        return path.resolve(strict=False)
    except (OSError, RuntimeError) as exc:
        raise InspectionError(f"{label}を安全に解決できません: {path}") from exc


def paths_refer_to_same_file(first: Path, second: Path) -> bool:
    if resolved_path(first, "入力パス") == resolved_path(second, "出力パス"):
        return True
    if not path_entry_exists(second):
        return False
    try:
        return os.path.samefile(first, second)
    except OSError:
        # 壊れたリンクなどは同一ファイルとはみなさない。既存出力の規則で扱う。
        return False


def validate_args(args: argparse.Namespace) -> Path:
    path = Path(args.workbook)
    if not path.exists():
        raise InspectionError(f"入力ファイルが存在しません: {path}")
    if path.is_symlink():
        raise InspectionError("シンボリックリンクのExcelファイルは対象外です。")
    if not path.is_file():
        raise InspectionError(f"通常ファイルではありません: {path}")
    suffix = path.suffix.casefold()
    if suffix not in SUPPORTED_SUFFIXES:
        if suffix == ".xls":
            raise InspectionError(".xlsは対象外です。.xlsxへ変換してから実行してください。")
        raise InspectionError("対応形式は.xlsxと.xlsmです。")
    for option in (
        "max_file_mb",
        "max_uncompressed_mb",
        "max_cells",
        "max_total_cells",
    ):
        if getattr(args, option) <= 0:
            raise InspectionError(f"--{option.replace('_', '-')}には1以上を指定してください。")
    size_limit = args.max_file_mb * 1024 * 1024
    if path.stat().st_size > size_limit:
        raise InspectionError(
            f"入力ファイルが上限{args.max_file_mb}MiBを超えています。"
            "対象シートや安全性を確認してから上限を変更してください。"
        )
    if args.force and not args.output:
        raise InspectionError("--forceは--outputと併用してください。")
    if args.output:
        destination = Path(args.output)
        if destination.suffix.casefold() != ".json":
            raise InspectionError("出力先の拡張子は.jsonだけを指定できます。")
        if paths_refer_to_same_file(path, destination):
            raise InspectionError(
                "出力先に入力ブックと同じファイルは指定できません。"
            )
        if destination.exists() and destination.is_dir():
            raise InspectionError(f"出力先がディレクトリです: {destination}")
        if path_entry_exists(destination) and not args.force:
            raise InspectionError(
                f"出力先が既に存在します: {destination}。"
                "上書きする場合だけ--forceを指定してください。"
            )
    return path


def xml_has_forbidden_declaration(member_stream: Any) -> bool:
    overlap = b""
    first_chunk = True
    looks_like_xml: bool | None = None
    while True:
        chunk = member_stream.read(64 * 1024)
        if not chunk:
            return False
        if first_chunk:
            first_chunk = False
            if chunk.startswith(b"\x4c\x6f\xa7\x94"):
                # XML標準のEBCDIC識別子。解析系ごとの差を避けるため拒否する。
                return True
        data = overlap + chunk
        normalized = data.replace(b"\x00", b"").upper()
        if looks_like_xml is None:
            prefix = normalized
            for marker in (b"\xef\xbb\xbf", b"\xff\xfe", b"\xfe\xff"):
                if prefix.startswith(marker.upper()):
                    prefix = prefix[len(marker) :]
                    break
            significant = prefix.lstrip(b" \t\r\n")
            if significant:
                looks_like_xml = significant.startswith(b"<")
                if not looks_like_xml:
                    return False
        if looks_like_xml and (
            b"<!DOCTYPE" in normalized or b"<!ENTITY" in normalized
        ):
            return True
        # UTF-16/32のコード単位と宣言文字列が境界をまたぐ場合に備える。
        overlap = data[-64:]


def content_type_is_xml(content_type: str) -> bool:
    media_type = content_type.split(";", 1)[0].strip().casefold()
    return (
        media_type.endswith("+xml")
        or media_type in {"application/xml", "text/xml"}
        or "vmldrawing" in media_type
        or media_type == "image/svg+xml"
    )


def relationship_target_path(
    relationships_name: str,
    target: str,
) -> str | None:
    decoded_target = unquote(target.split("#", 1)[0])
    if not decoded_target:
        return None
    if decoded_target.startswith("/"):
        normalized = posixpath.normpath(decoded_target).lstrip("/")
    else:
        relationships_path = PurePosixPath(relationships_name)
        source_directory = relationships_path.parent.parent
        normalized = posixpath.normpath(
            posixpath.join(str(source_directory), decoded_target)
        )
    if normalized == ".." or normalized.startswith("../"):
        raise InspectionError("ZIP外を指すRelationshipターゲットを検出しました。")
    return normalized


def xml_archive_members(
    archive: zipfile.ZipFile,
    infos: list[zipfile.ZipInfo],
) -> set[str]:
    content_types_info = next(
        info for info in infos if info.filename == "[Content_Types].xml"
    )
    if content_types_info.file_size > MAX_CONTENT_TYPES_BYTES:
        raise InspectionError("[Content_Types].xmlが安全上の上限を超えています。")
    with archive.open(content_types_info) as member_stream:
        if xml_has_forbidden_declaration(member_stream):
            raise InspectionError(
                "DTDまたはエンティティを含むか、対応外の符号化を使う"
                "XML部品は拒否しました。"
            )
    try:
        content_types_root = ElementTree.fromstring(
            archive.read(content_types_info)
        )
    except (ElementTree.ParseError, LookupError, UnicodeError) as exc:
        raise InspectionError("[Content_Types].xmlを安全に解析できません。") from exc

    defaults: dict[str, set[str]] = {}
    overrides: dict[str, set[str]] = {}
    for element in content_types_root:
        local_name = element.tag.rsplit("}", 1)[-1]
        content_type = element.attrib.get("ContentType", "")
        if local_name == "Default":
            extension = element.attrib.get("Extension", "").casefold()
            defaults.setdefault(extension, set()).add(content_type)
        elif local_name == "Override":
            part_name = element.attrib.get("PartName", "")
            if part_name and not part_name.startswith("/"):
                part_name = "/" + part_name
            overrides.setdefault(part_name, set()).add(content_type)

    candidates: set[str] = set()
    for info in infos:
        lower_name = info.filename.casefold()
        extension = PurePosixPath(lower_name).suffix.lstrip(".")
        part_name = "/" + info.filename.lstrip("/")
        declared_types = defaults.get(extension, set()) | overrides.get(
            part_name,
            set(),
        )
        if (
            lower_name.endswith((".xml", ".rels", ".vml"))
            or any(content_type_is_xml(item) for item in declared_types)
        ):
            candidates.add(info.filename)

    names = {info.filename for info in infos}
    relationship_infos = [
        info
        for info in infos
        if info.filename.casefold().endswith(".rels")
    ]
    for info in relationship_infos:
        if info.file_size > MAX_RELATIONSHIPS_BYTES:
            raise InspectionError("Relationship部品が安全上の上限を超えています。")
        with archive.open(info) as member_stream:
            if xml_has_forbidden_declaration(member_stream):
                raise InspectionError(
                    "DTDまたはエンティティを含むか、対応外の符号化を使う"
                    "XML部品は拒否しました。"
                )
        try:
            relationships_root = ElementTree.fromstring(archive.read(info))
        except (ElementTree.ParseError, LookupError, UnicodeError) as exc:
            raise InspectionError(
                f"Relationship部品を安全に解析できません: {info.filename}"
            ) from exc
        for relationship in relationships_root:
            if relationship.tag.rsplit("}", 1)[-1] != "Relationship":
                continue
            if relationship.attrib.get("TargetMode", "").casefold() == "external":
                continue
            target_name = relationship_target_path(
                info.filename,
                relationship.attrib.get("Target", ""),
            )
            if target_name in names:
                candidates.add(target_name)
    return candidates


def archive_metadata(path: Path, max_uncompressed_mb: int) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_PARTS:
                raise InspectionError(
                    f"ZIP部品数が上限{MAX_ARCHIVE_PARTS:,}件を超えています。"
                )

            total_uncompressed = sum(info.file_size for info in infos)
            limit = max_uncompressed_mb * 1024 * 1024
            if total_uncompressed > limit:
                raise InspectionError(
                    f"ZIP展開後の合計サイズが上限{max_uncompressed_mb}MiBを超えています。"
                )

            names = {info.filename for info in infos}
            if len(names) != len(infos):
                raise InspectionError("重複するZIP部品名を検出しました。")
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                raise InspectionError("Excelブックに必須のOOXML部品がありません。")

            for info in infos:
                member = PurePosixPath(info.filename)
                if member.is_absolute() or ".." in member.parts or "\\" in info.filename:
                    raise InspectionError("安全でないZIP部品パスを検出しました。")
                if info.flag_bits & 0x1:
                    raise InspectionError("暗号化されたZIP部品には対応していません。")
                if info.file_size > 1024 * 1024:
                    ratio = info.file_size / max(info.compress_size, 1)
                    if ratio > MAX_SUSPICIOUS_COMPRESSION_RATIO:
                        raise InspectionError("異常に高いZIP圧縮率を検出しました。")

            xml_members = xml_archive_members(archive, infos)
            for info in infos:
                if info.filename in xml_members:
                    with archive.open(info) as member_stream:
                        if xml_has_forbidden_declaration(member_stream):
                            raise InspectionError(
                                "DTDまたはエンティティを含むか、対応外の符号化を使う"
                                "XML部品は拒否しました。"
                            )

            folded_names = {name.casefold() for name in names}
            return {
                "uncompressed_size_bytes": total_uncompressed,
                "part_count": len(infos),
                "has_vba_project": "xl/vbaproject.bin" in folded_names,
                "external_link_parts": len(
                    [
                        name
                        for name in folded_names
                        if name.startswith("xl/externallinks/externallink")
                        and name.endswith(".xml")
                    ]
                ),
                "drawing_parts": len(
                    [
                        name
                        for name in folded_names
                        if name.startswith("xl/drawings/drawing")
                        and name.endswith(".xml")
                    ]
                ),
                "chart_parts": len(
                    [
                        name
                        for name in folded_names
                        if (
                            name.startswith("xl/charts/chart")
                            or "/charts/chart" in name
                        )
                        and name.endswith(".xml")
                    ]
                ),
                "media_parts": len(
                    [
                        name
                        for name in folded_names
                        if name.startswith("xl/media/")
                    ]
                ),
                "vml_drawing_parts": len(
                    [
                        name
                        for name in folded_names
                        if name.startswith("xl/drawings/vmldrawing")
                    ]
                ),
                "active_x_parts": len(
                    [
                        name
                        for name in folded_names
                        if name.startswith("xl/activex/")
                    ]
                ),
                "smartart_parts": len(
                    [
                        name
                        for name in folded_names
                        if name.startswith("xl/diagrams/")
                    ]
                ),
            }
    except (OSError, zipfile.BadZipFile) as exc:
        raise InspectionError(
            "有効なOffice Open XMLブックではありません。破損または暗号化を確認してください。"
        ) from exc


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return {"total_seconds": value.total_seconds()}
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def formula_payload(value: Any) -> Any:
    if isinstance(value, str):
        return value
    payload: dict[str, Any] = {"kind": value.__class__.__name__}
    for attribute in (
        "text",
        "ref",
        "ca",
        "dt2D",
        "dtr",
        "r1",
        "r2",
        "del1",
        "del2",
    ):
        if not hasattr(value, attribute):
            continue
        item = getattr(value, attribute)
        if item is not None:
            payload[attribute] = scalar(item)
    return payload


def color_value(color: Any) -> dict[str, Any] | None:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    if not color_type:
        return None
    value = getattr(color, color_type, None)
    if value is None:
        return None
    result: dict[str, Any] = {"type": color_type, "value": scalar(value)}
    tint = getattr(color, "tint", 0)
    if tint:
        result["tint"] = tint
    return result


def cell_style(cell: Any) -> dict[str, Any]:
    style: dict[str, Any] = {}
    font: dict[str, Any] = {}
    if cell.font.bold:
        font["bold"] = True
    if cell.font.italic:
        font["italic"] = True
    if cell.font.underline and cell.font.underline != "none":
        font["underline"] = cell.font.underline
    font_color = color_value(cell.font.color)
    if font_color:
        font["color"] = font_color
    if font:
        style["font"] = font

    if cell.fill.fill_type:
        style["fill"] = {
            "type": cell.fill.fill_type,
            "foreground": color_value(cell.fill.fgColor),
        }

    alignment: dict[str, Any] = {}
    for key in ("horizontal", "vertical", "wrap_text", "shrink_to_fit", "indent"):
        value = getattr(cell.alignment, key, None)
        if value not in (None, False, 0):
            alignment[key] = value
    if alignment:
        style["alignment"] = alignment

    if cell.style_id:
        style["style_id"] = cell.style_id
    return style


def anchor_text(obj: Any) -> str | None:
    anchor = getattr(obj, "anchor", None)
    if isinstance(anchor, str):
        return anchor
    start = getattr(anchor, "_from", None)
    if start is None:
        return None
    start_cell = f"{get_column_letter(start.col + 1)}{start.row + 1}"
    end = getattr(anchor, "to", None)
    if end is None:
        return start_cell
    end_cell = f"{get_column_letter(end.col + 1)}{end.row + 1}"
    return start_cell if start_cell == end_cell else f"{start_cell}:{end_cell}"


def object_inventory(objects: Iterable[Any], kind: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for obj in objects:
        item: dict[str, Any] = {
            "type": obj.__class__.__name__,
            "anchor": anchor_text(obj),
        }
        if kind == "image":
            image_format = getattr(obj, "format", None)
            if image_format:
                item["format"] = image_format
        result.append(item)
    return result


def table_inventory(ws: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for table in ws.tables.values():
        item: dict[str, Any] = {
            "name": table.name,
            "display_name": table.displayName,
            "ref": table.ref,
            "header_row_count": table.headerRowCount,
            "totals_row_count": table.totalsRowCount,
        }
        style = table.tableStyleInfo
        if style is not None:
            item["style"] = {
                "name": style.name,
                "show_first_column": style.showFirstColumn,
                "show_last_column": style.showLastColumn,
                "show_row_stripes": style.showRowStripes,
                "show_column_stripes": style.showColumnStripes,
            }
        result.append(item)
    return result


def interval_is_fully_hidden(
    start: int,
    end: int,
    hidden_indices: set[int],
) -> bool:
    interval_size = end - start + 1
    hidden_count = sum(start <= index <= end for index in hidden_indices)
    return hidden_count == interval_size


def validation_has_visible_target(
    validation: Any,
    hidden_rows: set[int],
    hidden_columns: set[int],
) -> bool:
    ranges = getattr(getattr(validation, "sqref", None), "ranges", ())
    for target in ranges:
        all_rows_hidden = interval_is_fully_hidden(
            target.min_row,
            target.max_row,
            hidden_rows,
        )
        all_columns_hidden = interval_is_fully_hidden(
            target.min_col,
            target.max_col,
            hidden_columns,
        )
        if not all_rows_hidden and not all_columns_hidden:
            return True
    return False


def validation_inventory(
    ws: Any,
    include_details: bool,
    include_hidden_cells: bool,
    hidden_rows: set[int],
    hidden_columns: set[int],
) -> list[dict[str, Any]]:
    container = getattr(ws, "data_validations", None)
    validations = getattr(container, "dataValidation", []) if container else []
    result: list[dict[str, Any]] = []
    for validation in validations:
        item: dict[str, Any] = {
            "ranges": str(validation.sqref),
            "type": validation.type,
            "operator": validation.operator,
            "allow_blank": validation.allowBlank,
        }
        show_details = include_details and (
            include_hidden_cells
            or validation_has_visible_target(
                validation,
                hidden_rows,
                hidden_columns,
            )
        )
        if show_details:
            item.update(
                {
                    "formula1": scalar(validation.formula1),
                    "formula2": scalar(validation.formula2),
                    "error_title": validation.errorTitle,
                    "error": validation.error,
                }
            )
        else:
            item["details_excluded"] = True
        result.append(item)
    return result


def references_excluded_sheet(
    value: Any,
    excluded_sheet_names: set[str],
) -> bool:
    if not isinstance(value, str):
        return False
    folded_value = value.casefold()
    for sheet_name in excluded_sheet_names:
        folded_name = sheet_name.casefold()
        escaped_name = sheet_name.replace("'", "''").casefold()
        # ローカル参照と外部ブック修飾参照の両方を保守的に検出する。
        if f"{folded_name}!" in folded_value:
            return True
        if f"{escaped_name}'!" in folded_value:
            return True
    return False


def defined_name_item(
    name: str,
    definition: Any,
    *,
    scope: str,
    sheet_name: str | None,
    excluded_sheet_names: set[str],
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "name": name,
        "scope": scope,
        "sheet": sheet_name,
        "type": scalar(getattr(definition, "type", None)),
        "hidden": bool(getattr(definition, "hidden", False)),
    }
    value = scalar(getattr(definition, "attr_text", None))
    scoped_to_excluded_sheet = (
        sheet_name is not None and sheet_name in excluded_sheet_names
    )
    if (
        item["hidden"]
        or scoped_to_excluded_sheet
        or references_excluded_sheet(value, excluded_sheet_names)
    ):
        item["value_excluded"] = True
    else:
        item["value"] = value
    return item


def defined_names(wb: Any, excluded_sheet_names: set[str]) -> list[dict[str, Any]]:
    result = [
        defined_name_item(
            name,
            definition,
            scope="workbook",
            sheet_name=None,
            excluded_sheet_names=excluded_sheet_names,
        )
        for name, definition in wb.defined_names.items()
    ]
    for ws in wb.worksheets:
        result.extend(
            defined_name_item(
                name,
                definition,
                scope="worksheet",
                sheet_name=ws.title,
                excluded_sheet_names=excluded_sheet_names,
            )
            for name, definition in ws.defined_names.items()
        )
    return result


def calculation_properties(wb: Any) -> dict[str, Any]:
    calculation = getattr(wb, "calculation", None)
    if calculation is None:
        return {}
    result: dict[str, Any] = {}
    for source, target in (
        ("calcMode", "mode"),
        ("fullCalcOnLoad", "full_calc_on_load"),
        ("forceFullCalc", "force_full_calc"),
        ("calcOnSave", "calc_on_save"),
        ("calcId", "calculation_id"),
    ):
        value = getattr(calculation, source, None)
        if value is not None:
            result[target] = scalar(value)
    return result


def declared_bounds(ws: Any) -> tuple[int, int, int, int, int]:
    min_row = max(int(ws.min_row or 1), 1)
    max_row = max(int(ws.max_row or 1), min_row)
    min_col = max(int(ws.min_column or 1), 1)
    max_col = max(int(ws.max_column or 1), min_col)
    grid_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    return min_row, max_row, min_col, max_col, grid_cells


def stored_meaningful_cells(ws: Any) -> list[Any]:
    """内容のある保存済みセルだけを取得し、膨張した宣言範囲の走査を避ける。"""
    cell_store = getattr(ws, "_cells", None)
    if not isinstance(cell_store, dict):
        raise InspectionError(
            f"シート「{ws.title}」のセル格納形式に対応していません。"
        )
    cells = []
    for cell in cell_store.values():
        if isinstance(cell, MergedCell):
            continue
        explicit_empty_string = cell.data_type in {"inlineStr", "s", "str"}
        if (
            cell.value is None
            and cell.comment is None
            and cell.hyperlink is None
            and not explicit_empty_string
        ):
            continue
        cells.append(cell)
    return sorted(cells, key=lambda item: (item.row, item.column))


def cell_is_hidden(ws: Any, cell: Any, hidden_columns: set[int]) -> bool:
    row_dimension = ws.row_dimensions.get(cell.row)
    if row_dimension is not None and row_dimension.hidden:
        return True
    return cell.column in hidden_columns


def hidden_column_indices(ws: Any) -> set[int]:
    result: set[int] = set()
    for key, dimension in ws.column_dimensions.items():
        if not dimension.hidden:
            continue
        start = dimension.min or column_index_from_string(key)
        end = dimension.max or start
        result.update(range(start, end + 1))
    return result


def effective_range(cells: list[Any]) -> str | None:
    if not cells:
        return None
    min_row = min(cell.row for cell in cells)
    max_row = max(cell.row for cell in cells)
    min_col = min(cell.column for cell in cells)
    max_col = max(cell.column for cell in cells)
    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def merged_range_map(
    ws: Any,
    include_values: bool,
    include_hidden_cells: bool,
    hidden_columns: set[int],
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    inventory: list[dict[str, Any]] = []
    anchors: dict[str, str] = {}
    merged_ranges = sorted(
        ws.merged_cells.ranges,
        key=lambda item: (
            item.min_row,
            item.min_col,
            item.max_row,
            item.max_col,
        ),
    )
    for merged in merged_ranges:
        anchor = ws.cell(merged.min_row, merged.min_col)
        range_text = str(merged)
        anchors[anchor.coordinate] = range_text
        item: dict[str, Any] = {
            "range": range_text,
            "anchor": anchor.coordinate,
        }
        if include_values and (
            include_hidden_cells
            or not cell_is_hidden(ws, anchor, hidden_columns)
        ):
            item["value"] = scalar(anchor.value)
        else:
            item["value_excluded"] = True
        inventory.append(item)
    return inventory, anchors


def make_cell_record(
    cell: Any,
    cached_ws: Any,
    merged_anchors: dict[str, str],
    include_styles: bool,
) -> dict[str, Any]:
    record: dict[str, Any] = {
        "coordinate": cell.coordinate,
        "row": cell.row,
        "column": get_column_letter(cell.column),
        "data_type": cell.data_type,
    }
    if cell.data_type == "f":
        record["formula"] = formula_payload(cell.value)
        cached_value = cached_ws[cell.coordinate].value
        record["cached_value"] = scalar(cached_value)
        record["cached_value_state"] = (
            "present" if cached_value is not None else "missing_or_empty"
        )
    else:
        record["value"] = (
            ""
            if cell.value is None and cell.data_type in {"inlineStr", "s", "str"}
            else scalar(cell.value)
        )

    if cell.coordinate in merged_anchors:
        record["merged_range"] = merged_anchors[cell.coordinate]
    if cell.comment is not None:
        record["comment"] = {
            "author": cell.comment.author,
            "text": cell.comment.text,
        }
    if cell.hyperlink is not None:
        record["hyperlink"] = {
            "target": cell.hyperlink.target,
            "location": cell.hyperlink.location,
            "display": cell.hyperlink.display,
        }
    if include_styles:
        style = cell_style(cell)
        if style:
            record["style"] = style
    if cell.number_format and cell.number_format != "General":
        record["number_format"] = cell.number_format
    return record


def row_blocks(row_columns: dict[int, list[int]]) -> list[dict[str, Any]]:
    if not row_columns:
        return []
    rows = sorted(row_columns)
    groups: list[list[int]] = [[rows[0]]]
    for row in rows[1:]:
        if row == groups[-1][-1] + 1:
            groups[-1].append(row)
        else:
            groups.append([row])

    result: list[dict[str, Any]] = []
    for group in groups:
        columns = [column for row in group for column in row_columns[row]]
        min_col = min(columns)
        max_col = max(columns)
        result.append(
            {
                "range": (
                    f"{get_column_letter(min_col)}{group[0]}:"
                    f"{get_column_letter(max_col)}{group[-1]}"
                ),
                "start_row": group[0],
                "end_row": group[-1],
                "min_column": get_column_letter(min_col),
                "max_column": get_column_letter(max_col),
            }
        )
    return result


def inspect_sheet(
    ws: Any,
    cached_ws: Any,
    included: bool,
    include_hidden_cells: bool,
    include_styles: bool,
    max_cells: int,
) -> dict[str, Any]:
    min_row, max_row, min_col, max_col, declared_grid_cells = declared_bounds(ws)
    hidden_rows = {
        row
        for row, dimension in ws.row_dimensions.items()
        if dimension.hidden
    }
    hidden_columns = hidden_column_indices(ws)
    if included:
        all_cells = stored_meaningful_cells(ws)
        extracted_cells = (
            all_cells
            if include_hidden_cells
            else [
                cell
                for cell in all_cells
                if not cell_is_hidden(ws, cell, hidden_columns)
            ]
        )
        excluded_hidden_cells = len(all_cells) - len(extracted_cells)
    else:
        all_cells = []
        extracted_cells = []
        excluded_hidden_cells = 0
    merged_ranges, merged_anchors = merged_range_map(
        ws,
        include_values=included,
        include_hidden_cells=include_hidden_cells,
        hidden_columns=hidden_columns,
    )
    warnings: list[str] = []
    charts = object_inventory(getattr(ws, "_charts", []), "chart")
    images = object_inventory(getattr(ws, "_images", []), "image")
    if charts:
        warnings.append("グラフの内容はJSONだけでは確認できません。シートを視覚確認してください。")
    if images:
        warnings.append("画像の内容はJSONへ抽出していません。シートを視覚確認してください。")
    if not included:
        warnings.append("このシートのセル内容は抽出対象外です。")
    elif excluded_hidden_cells:
        warnings.append(
            f"非表示行・列にある内容セル{excluded_hidden_cells}件を抽出対象外としました。"
        )

    result: dict[str, Any] = {
        "name": ws.title,
        "state": ws.sheet_state,
        "included": included,
        "dimensions": {
            "declared_range": ws.calculate_dimension(),
            "declared_min_row": min_row,
            "declared_max_row": max_row,
            "declared_min_column": get_column_letter(min_col),
            "declared_max_column": get_column_letter(max_col),
            "declared_grid_cells": declared_grid_cells,
            "effective_range": effective_range(all_cells) if included else None,
            "extracted_range": effective_range(extracted_cells) if included else None,
            "stored_content_cells": len(all_cells) if included else None,
            "extracted_content_cells": len(extracted_cells) if included else 0,
            "excluded_hidden_cells": excluded_hidden_cells if included else 0,
        },
        "freeze_panes": scalar(ws.freeze_panes),
        "auto_filter": scalar(getattr(ws.auto_filter, "ref", None)),
        "print_area": scalar(getattr(ws, "print_area", None)),
        "merged_ranges": merged_ranges,
        "tables": table_inventory(ws),
        "data_validations": validation_inventory(
            ws,
            include_details=included,
            include_hidden_cells=include_hidden_cells,
            hidden_rows=hidden_rows,
            hidden_columns=hidden_columns,
        ),
        "hidden_rows": [
            row
            for row in sorted(hidden_rows)
            if min_row <= row <= max_row
        ],
        "hidden_columns": [
            get_column_letter(column) for column in sorted(hidden_columns)
        ],
        "charts": charts,
        "images": images,
        "row_blocks": [],
        "rows": [],
        "warnings": warnings,
    }
    if not included:
        return result
    if len(extracted_cells) > max_cells:
        raise InspectionError(
            f"シート「{ws.title}」の内容セルが{len(extracted_cells):,}件で、"
            f"上限{max_cells:,}件を超えています。"
            "--sheetで対象を絞るか、上限変更の必要性を確認してください。"
        )

    rows_by_number: dict[int, list[dict[str, Any]]] = {}
    row_columns: dict[int, list[int]] = {}
    formula_count = 0
    for cell in extracted_cells:
        rows_by_number.setdefault(cell.row, []).append(
            make_cell_record(
                cell,
                cached_ws,
                merged_anchors,
                include_styles,
            )
        )
        row_columns.setdefault(cell.row, []).append(cell.column)
        if cell.data_type == "f":
            formula_count += 1

    rows_output = [
        {"row": row_number, "cells": rows_by_number[row_number]}
        for row_number in sorted(rows_by_number)
    ]

    result["rows"] = rows_output
    result["row_blocks"] = row_blocks(row_columns)
    result["dimensions"]["formula_cells"] = formula_count
    if formula_count:
        warnings.append(
            "数式のcached_valueはExcelで最後に保存された値であり、欠落または古い場合があります。"
        )
    return result


def selected_sheet_names(wb: Any, args: argparse.Namespace) -> set[str]:
    worksheet_names = {ws.title for ws in wb.worksheets}
    chartsheet_names = {
        sheet.title for sheet in getattr(wb, "chartsheets", [])
    }
    available = worksheet_names | chartsheet_names
    requested = set(args.sheet)
    missing = sorted(requested - available)
    if missing:
        raise InspectionError(
            "指定したシートが存在しません: " + ", ".join(missing)
        )
    unsupported = sorted(requested & chartsheet_names)
    if unsupported:
        raise InspectionError(
            "グラフシートはセル抽出対象外です: " + ", ".join(unsupported)
        )
    if requested:
        return requested
    return {
        ws.title
        for ws in wb.worksheets
        if args.include_hidden_sheets or ws.sheet_state == "visible"
    }


def chartsheet_inventory(wb: Any) -> list[dict[str, Any]]:
    return [
        {
            "name": sheet.title,
            "state": sheet.sheet_state,
        }
        for sheet in getattr(wb, "chartsheets", [])
    ]


def inspect_workbook(path: Path, args: argparse.Namespace) -> dict[str, Any]:
    archive = archive_metadata(path, args.max_uncompressed_mb)
    macro_enabled = path.suffix.casefold() == ".xlsm"
    try:
        formula_wb = openpyxl.load_workbook(
            path,
            read_only=False,
            data_only=False,
            keep_vba=False,
            keep_links=False,
        )
        cached_wb = openpyxl.load_workbook(
            path,
            read_only=False,
            data_only=True,
            keep_vba=False,
            keep_links=False,
        )
    except Exception as exc:
        raise InspectionError(
            f"Excelブックを読み取れません: {exc.__class__.__name__}: {exc}"
        ) from exc

    selected = selected_sheet_names(formula_wb, args)
    total_cells = 0
    for ws in formula_wb.worksheets:
        if ws.title not in selected:
            continue
        cells = stored_meaningful_cells(ws)
        if not args.include_hidden_rows_columns:
            hidden_columns = hidden_column_indices(ws)
            cells = [
                cell
                for cell in cells
                if not cell_is_hidden(ws, cell, hidden_columns)
            ]
        total_cells += len(cells)
    if total_cells > args.max_total_cells:
        raise InspectionError(
            f"対象シート合計の内容セルが{total_cells:,}件で、"
            f"上限{args.max_total_cells:,}件を超えています。"
            "--sheetで対象を絞ってください。"
        )

    sheets: list[dict[str, Any]] = []
    for ws in formula_wb.worksheets:
        sheets.append(
            inspect_sheet(
                ws,
                cached_wb[ws.title],
                included=ws.title in selected,
                include_hidden_cells=args.include_hidden_rows_columns,
                include_styles=not args.no_styles,
                max_cells=args.max_cells,
            )
        )

    warnings = [
        "マクロと数式は実行していません。",
        "数式のcached_valueは最新であることを保証しません。",
    ]
    if macro_enabled or archive["has_vba_project"]:
        warnings.append("VBAは読み込み、解析、実行のいずれもしていません。")
    if archive["external_link_parts"]:
        warnings.append(
            f"外部リンク定義{archive['external_link_parts']}件を検出しましたが、参照先へアクセスしていません。"
        )
    object_parts = (
        archive["drawing_parts"]
        + archive["chart_parts"]
        + archive["media_parts"]
        + archive["vml_drawing_parts"]
        + archive["active_x_parts"]
        + archive["smartart_parts"]
    )
    if object_parts:
        warnings.append(
            "描画部品を検出しました。JSONのグラフ・画像一覧は種類と位置だけで、"
            "図形、SmartArt、ActiveX、VMLおよびグラフ系列の詳細を完全には解析しません。"
        )
    chartsheets = chartsheet_inventory(formula_wb)
    if chartsheets:
        warnings.append(
            "グラフシートは一覧だけを記録し、内容を抽出していません: "
            + ", ".join(sheet["name"] for sheet in chartsheets)
        )
    excluded = [sheet["name"] for sheet in sheets if not sheet["included"]]
    if excluded:
        warnings.append("セル内容を抽出しなかったシート: " + ", ".join(excluded))
    has_included_formulas = any(
        sheet["included"]
        and sheet["dimensions"].get("formula_cells", 0)
        for sheet in sheets
    )
    has_excluded_hidden_cells = any(
        sheet["included"]
        and sheet["dimensions"].get("excluded_hidden_cells", 0)
        for sheet in sheets
    )
    if has_included_formulas and (excluded or has_excluded_hidden_cells):
        warnings.append(
            "表示セルのcached_valueが除外したシート・行・列に由来するかは"
            "追跡していません。派生値も除外する用途では使用しないでください。"
        )

    return {
        "schema_version": SCHEMA_VERSION,
        "source": {
            "file_name": path.name,
            "suffix": path.suffix.casefold(),
            "size_bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        },
        "workbook": {
            "openpyxl_version": openpyxl.__version__,
            "archive": archive,
            "date_system": scalar(getattr(formula_wb, "epoch", None)),
            "sheet_count": len(formula_wb.sheetnames),
            "worksheet_count": len(formula_wb.worksheets),
            "chartsheet_count": len(chartsheets),
            "included_sheet_count": len(selected),
            "sheet_order": list(formula_wb.sheetnames),
            "chartsheets": chartsheets,
            "defined_names": defined_names(formula_wb, set(excluded)),
            "calculation": calculation_properties(formula_wb),
        },
        "sheets": sheets,
        "warnings": warnings,
    }


def write_json(
    payload: dict[str, Any],
    output: str | None,
    *,
    force: bool,
) -> None:
    text = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False) + "\n"
    if output is None:
        sys.stdout.write(text)
        return
    destination = Path(output)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if path_entry_exists(destination) and not force:
        raise InspectionError(
            f"出力先が既に存在します: {destination}。"
            "上書きする場合だけ--forceを指定してください。"
        )
    descriptor, temporary = tempfile.mkstemp(
        prefix=f".{destination.name}.",
        suffix=".tmp",
        dir=destination.parent,
        text=True,
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as stream:
            stream.write(text)
            stream.flush()
            os.fsync(stream.fileno())
        if force:
            os.replace(temporary, destination)
        else:
            try:
                # 完成済み一時ファイルを、既存パスを置換せずatomicに公開する。
                os.link(temporary, destination)
            except FileExistsError as exc:
                raise InspectionError(
                    f"出力先が既に存在します: {destination}。"
                    "上書きする場合だけ--forceを指定してください。"
                ) from exc
            os.unlink(temporary)
    except Exception:
        try:
            os.unlink(temporary)
        except OSError:
            pass
        raise


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    try:
        path = validate_args(args)
        payload = inspect_workbook(path, args)
        write_json(payload, args.output, force=args.force)
    except InspectionError as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 2
    except OSError as exc:
        print(f"エラー: 出力に失敗しました: {exc}", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
