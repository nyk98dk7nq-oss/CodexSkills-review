from __future__ import annotations

import json
import hashlib
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo


SKILL_DIR = Path(__file__).resolve().parents[1]
SCRIPT = SKILL_DIR / "scripts" / "inspect_excel.py"


class InspectExcelTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.directory = Path(self.temporary.name)
        self.workbook_path = self.directory / "sample.xlsx"
        self._build_workbook(self.workbook_path)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    @staticmethod
    def _build_workbook(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "要件一覧"
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "要件一覧"
        sheet["A2"] = "ID"
        sheet["B2"] = "観点"
        sheet["C2"] = "値"
        sheet["A3"] = "001"
        sheet["B3"] = "確認A\n確認B"
        sheet["B3"].comment = Comment("二つの観点を分離する", "Tester")
        sheet["C3"] = "=SUM(1,1)"
        sheet["A4"] = "002"
        sheet["B4"] = "非表示行の内容"
        sheet["B4"].hyperlink = "https://example.com/reference"
        sheet["C4"] = 0
        sheet.row_dimensions[4].hidden = True
        sheet["D3"] = "非表示列の内容"
        sheet.column_dimensions["D"].hidden = True
        sheet["E3"] = ArrayFormula(ref="E3:E4", text="=ROW(E3:E4)")
        sheet["G3"] = ""
        sheet.merge_cells("F4:G4")
        sheet["F4"] = "ROW-SECRET"

        table = Table(displayName="Requirements", ref="A2:C4")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        validation = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
        validation.add("C3:C4")
        sheet.add_data_validation(validation)
        hidden_validation = DataValidation(
            type="list",
            formula1='"VALIDATION-SECRET"',
            errorTitle="HIDDEN-ERROR-TITLE",
            error="HIDDEN-ERROR",
        )
        hidden_validation.add("H4")
        sheet.add_data_validation(hidden_validation)

        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=3, min_row=2, max_row=4), titles_from_data=True)
        sheet.add_chart(chart, "F2")

        hidden = workbook.create_sheet("非表示")
        hidden.merge_cells("A1:B1")
        hidden["A1"] = "TOP-SECRET"
        hidden.sheet_state = "hidden"

        workbook.defined_names.add(
            DefinedName("WorkbookVisible", attr_text="'要件一覧'!$A$1")
        )
        workbook.defined_names.add(
            DefinedName("WorkbookHiddenRef", attr_text="'非表示'!$A$1")
        )
        sheet.defined_names.add(
            DefinedName("VisibleLocal", attr_text="'要件一覧'!$A$2")
        )
        sheet.defined_names.add(
            DefinedName("VisibleLocalHiddenRef", attr_text="'非表示'!$A$1")
        )
        hidden.defined_names.add(
            DefinedName("HiddenLocal", attr_text='"LOCAL-SHEET-SECRET"')
        )
        workbook.save(path)

    def run_script(self, *arguments: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(SCRIPT), str(self.workbook_path), *arguments],
            check=False,
            capture_output=True,
            text=True,
        )

    def test_default_extracts_visible_content_and_inventories_hidden_elements(self) -> None:
        before_hash = hashlib.sha256(self.workbook_path.read_bytes()).hexdigest()
        completed = self.run_script()
        self.assertEqual(completed.returncode, 0, completed.stderr)
        payload = json.loads(completed.stdout)

        self.assertEqual(payload["schema_version"], "1.1")
        self.assertEqual(payload["source"]["sha256"], before_hash)
        self.assertEqual(
            hashlib.sha256(self.workbook_path.read_bytes()).hexdigest(),
            before_hash,
        )
        self.assertEqual(payload["workbook"]["included_sheet_count"], 1)
        self.assertEqual(payload["workbook"]["archive"]["chart_parts"], 1)
        self.assertEqual(payload["workbook"]["archive"]["drawing_parts"], 1)
        self.assertTrue(
            any("描画部品" in warning for warning in payload["warnings"])
        )
        sheet = next(item for item in payload["sheets"] if item["name"] == "要件一覧")
        coordinates = {
            cell["coordinate"]
            for row in sheet["rows"]
            for cell in row["cells"]
        }
        self.assertIn("A1", coordinates)
        self.assertIn("C3", coordinates)
        self.assertIn("E3", coordinates)
        self.assertIn("G3", coordinates)
        self.assertNotIn("A4", coordinates)
        self.assertNotIn("D3", coordinates)
        self.assertEqual(sheet["hidden_rows"], [4])
        self.assertIn("D", sheet["hidden_columns"])
        self.assertEqual(sheet["merged_ranges"][0]["range"], "A1:C1")
        hidden_row_merge = next(
            item for item in sheet["merged_ranges"] if item["range"] == "F4:G4"
        )
        self.assertEqual(hidden_row_merge["value_excluded"], True)
        self.assertNotIn("value", hidden_row_merge)
        self.assertNotIn("ROW-SECRET", completed.stdout)
        self.assertEqual(sheet["tables"][0]["name"], "Requirements")
        self.assertEqual(sheet["data_validations"][0]["ranges"], "C3:C4")
        redacted_validation = next(
            item for item in sheet["data_validations"] if item["ranges"] == "H4"
        )
        self.assertEqual(redacted_validation["details_excluded"], True)
        self.assertNotIn("formula1", redacted_validation)
        self.assertNotIn("VALIDATION-SECRET", completed.stdout)
        self.assertNotIn("HIDDEN-ERROR", completed.stdout)
        self.assertEqual(len(sheet["charts"]), 1)
        formula = next(
            cell
            for row in sheet["rows"]
            for cell in row["cells"]
            if cell["coordinate"] == "C3"
        )
        self.assertEqual(formula["formula"], "=SUM(1,1)")
        self.assertIsNone(formula["cached_value"])
        self.assertEqual(formula["cached_value_state"], "missing_or_empty")
        array_formula = next(
            cell
            for row in sheet["rows"]
            for cell in row["cells"]
            if cell["coordinate"] == "E3"
        )
        self.assertEqual(array_formula["formula"]["kind"], "ArrayFormula")
        self.assertEqual(array_formula["formula"]["ref"], "E3:E4")
        self.assertEqual(array_formula["formula"]["text"], "=ROW(E3:E4)")
        empty_string = next(
            cell
            for row in sheet["rows"]
            for cell in row["cells"]
            if cell["coordinate"] == "G3"
        )
        self.assertEqual(empty_string["value"], "")

        hidden = next(item for item in payload["sheets"] if item["name"] == "非表示")
        self.assertFalse(hidden["included"])
        self.assertEqual(hidden["rows"], [])
        self.assertEqual(hidden["merged_ranges"][0]["value_excluded"], True)
        self.assertNotIn("value", hidden["merged_ranges"][0])
        self.assertNotIn("TOP-SECRET", completed.stdout)

    def test_include_hidden_sheets_does_not_include_hidden_rows_or_columns(self) -> None:
        completed = self.run_script("--include-hidden-sheets")
        self.assertEqual(completed.returncode, 0, completed.stderr)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["workbook"]["included_sheet_count"], 2)

        sheet = next(item for item in payload["sheets"] if item["name"] == "要件一覧")
        coordinates = {
            cell["coordinate"]
            for row in sheet["rows"]
            for cell in row["cells"]
        }
        self.assertNotIn("A4", coordinates)
        self.assertNotIn("D3", coordinates)
        hidden = next(item for item in payload["sheets"] if item["name"] == "非表示")
        self.assertTrue(hidden["included"])
        self.assertEqual(hidden["rows"][0]["cells"][0]["value"], "TOP-SECRET")

    def test_include_hidden_rows_columns_does_not_include_hidden_sheets(self) -> None:
        completed = self.run_script("--include-hidden-rows-columns")
        self.assertEqual(completed.returncode, 0, completed.stderr)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["workbook"]["included_sheet_count"], 1)

        sheet = next(item for item in payload["sheets"] if item["name"] == "要件一覧")
        coordinates = {
            cell["coordinate"]
            for row in sheet["rows"]
            for cell in row["cells"]
        }
        self.assertIn("A4", coordinates)
        self.assertIn("D3", coordinates)
        self.assertIn("F4", coordinates)
        hidden_validation = next(
            item for item in sheet["data_validations"] if item["ranges"] == "H4"
        )
        self.assertEqual(hidden_validation["formula1"], '"VALIDATION-SECRET"')
        hidden = next(item for item in payload["sheets"] if item["name"] == "非表示")
        self.assertFalse(hidden["included"])
        self.assertNotIn("TOP-SECRET", completed.stdout)

    def test_explicit_sheet_can_select_one_hidden_sheet(self) -> None:
        completed = self.run_script("--sheet", "非表示")
        self.assertEqual(completed.returncode, 0, completed.stderr)
        payload = json.loads(completed.stdout)
        self.assertEqual(payload["workbook"]["included_sheet_count"], 1)

        visible = next(item for item in payload["sheets"] if item["name"] == "要件一覧")
        self.assertFalse(visible["included"])
        self.assertIsNone(visible["dimensions"]["effective_range"])
        hidden = next(item for item in payload["sheets"] if item["name"] == "非表示")
        self.assertTrue(hidden["included"])
        self.assertEqual(hidden["rows"][0]["cells"][0]["value"], "TOP-SECRET")

    def test_output_is_deterministic(self) -> None:
        first = self.run_script("--no-styles")
        second = self.run_script("--no-styles")
        self.assertEqual(first.returncode, 0, first.stderr)
        self.assertEqual(second.returncode, 0, second.stderr)
        self.assertEqual(first.stdout, second.stdout)

    def test_defined_names_include_scope_and_redact_excluded_sheets(self) -> None:
        completed = self.run_script()
        self.assertEqual(completed.returncode, 0, completed.stderr)
        names = {
            (item["scope"], item["sheet"], item["name"]): item
            for item in json.loads(completed.stdout)["workbook"]["defined_names"]
        }

        workbook_visible = names[("workbook", None, "WorkbookVisible")]
        self.assertEqual(workbook_visible["value"], "'要件一覧'!$A$1")
        local_visible = names[("worksheet", "要件一覧", "VisibleLocal")]
        self.assertEqual(local_visible["value"], "'要件一覧'!$A$2")

        for key in (
            ("workbook", None, "WorkbookHiddenRef"),
            ("worksheet", "要件一覧", "VisibleLocalHiddenRef"),
            ("worksheet", "非表示", "HiddenLocal"),
        ):
            self.assertTrue(names[key]["value_excluded"])
            self.assertNotIn("value", names[key])
        self.assertNotIn("LOCAL-SHEET-SECRET", completed.stdout)

    def test_output_requires_json_extension(self) -> None:
        output = self.directory / "inspection.txt"
        completed = self.run_script("--output", str(output))
        self.assertEqual(completed.returncode, 2)
        self.assertIn("拡張子は.jsonだけ", completed.stderr)
        self.assertFalse(output.exists())

    def test_existing_output_requires_force_and_force_replaces_atomically(self) -> None:
        output = self.directory / "inspection.json"
        output.write_text("KEEP", encoding="utf-8")

        refused = self.run_script("--output", str(output))
        self.assertEqual(refused.returncode, 2)
        self.assertIn("--force", refused.stderr)
        self.assertEqual(output.read_text(encoding="utf-8"), "KEEP")

        replaced = self.run_script("--output", str(output), "--force")
        self.assertEqual(replaced.returncode, 0, replaced.stderr)
        self.assertEqual(json.loads(output.read_text(encoding="utf-8"))["schema_version"], "1.1")
        self.assertEqual(list(self.directory.glob(f".{output.name}.*.tmp")), [])

    def test_output_symlink_to_input_is_rejected_even_with_force(self) -> None:
        output = self.directory / "inspection.json"
        try:
            output.symlink_to(self.workbook_path)
        except OSError as exc:  # pragma: no cover - OSまたは権限依存
            self.skipTest(f"シンボリックリンクを作成できません: {exc}")
        before = self.workbook_path.read_bytes()

        completed = self.run_script("--output", str(output), "--force")
        self.assertEqual(completed.returncode, 2)
        self.assertIn("入力ブックと同じファイル", completed.stderr)
        self.assertEqual(self.workbook_path.read_bytes(), before)
        self.assertTrue(output.is_symlink())

    def test_force_replaces_output_symlink_without_following_it(self) -> None:
        target = self.directory / "symlink-target.txt"
        target.write_text("KEEP-TARGET", encoding="utf-8")
        output = self.directory / "inspection.json"
        try:
            output.symlink_to(target)
        except OSError as exc:  # pragma: no cover - OSまたは権限依存
            self.skipTest(f"シンボリックリンクを作成できません: {exc}")

        completed = self.run_script("--output", str(output), "--force")
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertFalse(output.is_symlink())
        self.assertEqual(target.read_text(encoding="utf-8"), "KEEP-TARGET")
        self.assertEqual(json.loads(output.read_text(encoding="utf-8"))["schema_version"], "1.1")

    def test_hard_link_to_input_is_rejected_even_with_force(self) -> None:
        output = self.directory / "inspection.json"
        try:
            output.hardlink_to(self.workbook_path)
        except OSError as exc:  # pragma: no cover - OSまたは権限依存
            self.skipTest(f"ハードリンクを作成できません: {exc}")
        before = self.workbook_path.read_bytes()

        completed = self.run_script("--output", str(output), "--force")
        self.assertEqual(completed.returncode, 2)
        self.assertIn("入力ブックと同じファイル", completed.stderr)
        self.assertEqual(self.workbook_path.read_bytes(), before)

    def test_openpyxl_3_2_is_rejected(self) -> None:
        code = (
            "import openpyxl, runpy; "
            "openpyxl.__version__ = '3.2.0'; "
            f"runpy.run_path({str(SCRIPT)!r}, run_name='version_check')"
        )
        completed = subprocess.run(
            [sys.executable, "-c", code],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(completed.returncode, 3)
        self.assertIn("openpyxl>=3.1,<3.2", completed.stderr)

    def test_cell_limit_fails_without_silent_truncation(self) -> None:
        completed = self.run_script("--max-cells", "1")
        self.assertEqual(completed.returncode, 2)
        self.assertIn("上限1件を超えています", completed.stderr)

    def test_legacy_xls_is_rejected(self) -> None:
        legacy = self.directory / "legacy.xls"
        legacy.write_bytes(b"not-an-xls")
        completed = subprocess.run(
            [sys.executable, str(SCRIPT), str(legacy)],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(completed.returncode, 2)
        self.assertIn(".xlsは対象外", completed.stderr)

    def test_invalid_ooxml_is_rejected(self) -> None:
        invalid = self.directory / "invalid.xlsx"
        invalid.write_bytes(b"not-a-zip")
        completed = subprocess.run(
            [sys.executable, str(SCRIPT), str(invalid)],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(completed.returncode, 2)
        self.assertIn("有効なOffice Open XML", completed.stderr)

    def test_late_dtd_declaration_is_rejected(self) -> None:
        malicious = self.directory / "late-dtd.xlsx"
        with zipfile.ZipFile(self.workbook_path) as source, zipfile.ZipFile(
            malicious, "w"
        ) as destination:
            for info in source.infolist():
                content = source.read(info)
                if info.filename == "xl/worksheets/sheet1.xml":
                    insertion = (
                        b" " * 70_000
                        + b'<!DOCTYPE worksheet [<!ENTITY hidden "blocked">]>\n'
                    )
                    content = insertion + content
                destination.writestr(info, content)

        completed = subprocess.run(
            [sys.executable, str(SCRIPT), str(malicious)],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(completed.returncode, 2)
        self.assertIn("DTDまたはエンティティ", completed.stderr)

    def test_uppercase_utf16_xml_with_late_dtd_is_rejected(self) -> None:
        malicious = self.directory / "uppercase-utf16-dtd.xlsx"
        source_sheet = "xl/worksheets/sheet1.xml"
        target_sheet = "xl/worksheets/sheet1.XML"
        with zipfile.ZipFile(self.workbook_path) as source, zipfile.ZipFile(
            malicious, "w"
        ) as destination:
            for info in source.infolist():
                target_name = (
                    target_sheet if info.filename == source_sheet else info.filename
                )
                content = source.read(info)
                if info.filename in {
                    "[Content_Types].xml",
                    "xl/_rels/workbook.xml.rels",
                }:
                    content = content.replace(b"sheet1.xml", b"sheet1.XML")
                if info.filename == source_sheet:
                    xml_text = content.decode("utf-8")
                    xml_text = xml_text.replace("要件一覧", "&hidden;", 1)
                    xml_text = (
                        '<?xml version="1.0" encoding="UTF-16"?>\n'
                        + " " * 70_000
                        + '<!DOCTYPE worksheet [<!ENTITY hidden "EXPANDED">]>\n'
                        + xml_text
                    )
                    content = xml_text.encode("utf-16")
                destination.writestr(
                    target_name,
                    content,
                    compress_type=info.compress_type,
                )

        completed = subprocess.run(
            [sys.executable, str(SCRIPT), str(malicious)],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(completed.returncode, 2)
        self.assertIn("DTDまたはエンティティ", completed.stderr)

    def test_internal_relationship_target_is_scanned_despite_spoofed_metadata(self) -> None:
        malicious = self.directory / "worksheet-dat-dtd.xlsx"
        source_sheet = "xl/worksheets/sheet1.xml"
        target_sheet = "xl/worksheets/sheet1.dat"
        with zipfile.ZipFile(self.workbook_path) as source, zipfile.ZipFile(
            malicious, "w"
        ) as destination:
            for info in source.infolist():
                target_name = (
                    target_sheet if info.filename == source_sheet else info.filename
                )
                content = source.read(info)
                if info.filename == "xl/_rels/workbook.xml.rels":
                    content = content.replace(b"sheet1.xml", b"sheet1.dat")
                    content = content.replace(b"/worksheet", b"/notxml", 1)
                if info.filename == source_sheet:
                    content = (
                        b" " * 70_000
                        + b'<!DOCTYPE worksheet [<!ENTITY hidden "EXPANDED">]>\n'
                        + content.replace(
                            "要件一覧".encode(),
                            b"&hidden;",
                            1,
                        )
                    )
                destination.writestr(
                    target_name,
                    content,
                    compress_type=info.compress_type,
                )

        completed = subprocess.run(
            [sys.executable, str(SCRIPT), str(malicious)],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(completed.returncode, 2)
        self.assertIn("DTDまたはエンティティ", completed.stderr)


if __name__ == "__main__":
    unittest.main()
